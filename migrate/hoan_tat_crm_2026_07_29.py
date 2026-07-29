"""Hoàn tất đồng bộ CRM (user duyệt 2026-07-29):
  1. Tạo khách mới (SĐT trong file Contact chưa có trong DB).
  2. Đổi tên khách theo file (các ca tên khác) — TRỪ ca tag đại lý khác (Hoa Vũ 24K/HAINAM).
  3. Thêm ticket mới (mã trong file Tickets chưa có trong DB), khớp serial->khách.

Khớp khách theo SĐT chuẩn hoá. Nguồn: File gốc/Hệ thống CRM/*.xlsx
Chạy:  .venv/bin/python -m migrate.hoan_tat_crm_2026_07_29 [--ghi]
"""
import datetime as dt
import json
import os
import pathlib
import sys
import urllib.parse
import urllib.request

import openpyxl
from migrate.parse import normalize_phone
from migrate.cap_nhat_khach_crm import clean_tinh

ROOT = pathlib.Path(__file__).resolve().parent.parent
CRM = ROOT / "File gốc/Hệ thống CRM"
# SĐT không đổi tên (ca tag đại lý khác cần hỏi riêng) — truyền qua env để KHÔNG hardcode
# thông tin khách vào code:  SKIP_RENAME="0xxxxxxxxx,0yyyyyyyyy" .venv/bin/python -m ...
SKIP_RENAME = set(filter(None, (s.strip() for s in os.environ.get("SKIP_RENAME", "").split(","))))


def dest():
    env = dict(l.strip().split("=", 1) for l in (ROOT / "migrate/.env.migrate").read_text().splitlines()
               if l.strip() and not l.startswith("#") and "=" in l)
    url, key = env["DEST_URL"], env["DEST_SERVICE_KEY"]
    assert "bwzmqfbcgouhvhoslmmm" in url
    return url, key


def rest(url, key, path, method="GET", body=None, prefer=None):
    p = urllib.parse.quote(path, safe='?&=,.*"')
    h = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}
    if prefer:
        h["Prefer"] = prefer
    req = urllib.request.Request(f"{url}/rest/v1/{p}", method=method,
                                 data=json.dumps(body).encode() if body is not None else None, headers=h)
    with urllib.request.urlopen(req) as r:
        raw = r.read()
        return json.loads(raw) if raw else None


def doc_het(url, key, path):
    rows, off = [], 0
    while True:
        b = rest(url, key, f"{path}{'&' if '?' in path else '?'}limit=1000&offset={off}")
        rows += b
        if len(b) < 1000:
            return rows
        off += 1000


def iso(v):
    if isinstance(v, (dt.datetime, dt.date)):
        return v.isoformat()
    return str(v).strip() if v else None


def main():
    ghi = "--ghi" in sys.argv
    url, key = dest()

    # ── đọc file khách ──
    wb = openpyxl.load_workbook(CRM / "Contact (res.partner).xlsx", read_only=True, data_only=True)
    ws = wb.active
    by_phone = {}
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r[0] or not r[2]:
            continue
        sdt, err = normalize_phone(r[2])
        if err:
            continue
        by_phone[sdt] = {"ten": str(r[0]).strip(),
                         "dia_chi": (str(r[3]).strip() if r[3] else None), "tinh": clean_tinh(r[4])}
    wb.close()

    db = doc_het(url, key, "cs_customers?select=id,primary_phone,full_name")
    db_by_phone = {c["primary_phone"]: c for c in db if c["primary_phone"]}

    # 1. khách mới
    moi = [{"full_name": f["ten"], "primary_phone": sdt, "address": f["dia_chi"],
            "province": f["tinh"], "source": "CRM Odoo", "needs_phone": False}
           for sdt, f in by_phone.items() if sdt not in db_by_phone]
    # 2. đổi tên
    doi_ten = [(db_by_phone[sdt]["id"], f["ten"], db_by_phone[sdt]["full_name"], sdt)
               for sdt, f in by_phone.items()
               if sdt in db_by_phone and f["ten"] and f["ten"] != db_by_phone[sdt]["full_name"]
               and sdt not in SKIP_RENAME]

    # ── 3. ticket mới ──
    wb = openpyxl.load_workbook(CRM / "Tickets (gwt.ticket).xlsx", read_only=True, data_only=True)
    ws = wb.active
    db_codes = {t["ticket_code"] for t in doc_het(url, key, "tickets?select=ticket_code")}
    ib = {m["serial"]: m["customer_id"] for m in doc_het(url, key, "installed_base?select=serial,customer_id")}
    seen = set()
    tickets = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        code = str(r[9]).strip() if r[9] else None
        if not code or code in db_codes or code in seen:
            continue
        seen.add(code)
        serial = str(r[4]).strip() if r[4] else None
        in_db = serial in ib if serial else False
        tickets.append({
            "ticket_code": code,
            "serial": serial if in_db else None,
            "source_serial": serial,
            "customer_id": ib.get(serial) if in_db else None,
            "source_customer": str(r[0]).strip() if r[0] else None,
            "ticket_type": str(r[1]).strip() if r[1] else None,
            "description": str(r[3]).strip() if r[3] else None,
            "last_note": str(r[6]).strip() if r[6] else None,
            "province": str(r[7]).strip() if r[7] else None,
            "state": str(r[8]).strip() if r[8] else "Open",
            "created_at": iso(r[2]),
        })
    wb.close()

    print(f"1. Khách mới tạo:      {len(moi)}")
    print(f"2. Đổi tên theo file:  {len(doi_ten)} (bỏ {len(SKIP_RENAME)} ca tag đại lý)")
    print(f"3. Ticket mới thêm:    {len(tickets)}  "
          f"({sum(1 for t in tickets if t['customer_id'])} khớp khách qua serial)")
    for t in tickets:
        print(f"     {t['ticket_code']} · serial {t['source_serial']} · "
              f"{'khớp khách' if t['customer_id'] else 'chưa khớp khách'} · {t['state']}")

    if not ghi:
        print("\n(dry-run — thêm --ghi để ghi thật)")
        return
    if moi:
        for i in range(0, len(moi), 100):
            rest(url, key, "cs_customers", method="POST", body=moi[i:i + 100], prefer="return=minimal")
    for cid, ten, _old, _sdt in doi_ten:
        rest(url, key, f"cs_customers?id=eq.{cid}", method="PATCH",
             body={"full_name": ten}, prefer="return=minimal")
    if tickets:
        rest(url, key, "tickets", method="POST", body=tickets, prefer="return=minimal")
    print(f"\nĐÃ GHI: {len(moi)} khách mới · {len(doi_ten)} đổi tên · {len(tickets)} ticket mới.")


if __name__ == "__main__":
    main()
