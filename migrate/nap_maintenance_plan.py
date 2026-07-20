"""Nạp gói bảo trì POE vào maintenance_plan.

Nguồn: file Excel user đã duyệt tay "GWT_goi_bao_tri_tu_hop_dong_2026-07-17 (1).xlsx"
(78 khách, cột LOẠI GÓI = HỢP ĐỒNG/TẶNG-NỘI BỘ/trống) + trích SĐT trực tiếp từ hợp đồng
để khớp customer_id (không khớp bằng tên — bài học từ ca "2 khách tên Yến").

Trích SĐT: chỉ tin số nằm trong khối "BÊN MUA/BÁN (A)" (= khách) tới trước "(B)" (= GWT),
loại trừ SĐT/MST công ty GWT đã biết (0339946388, 0110530659) — 2 nguồn nhiễu xác nhận
lặp lại ở nhiều hợp đồng khác nhau khi thử quét ẩu ban đầu.

Chạy:  .venv/bin/python -m migrate.nap_maintenance_plan          # dry-run
       .venv/bin/python -m migrate.nap_maintenance_plan --ghi    # ghi thật
"""

import json
import pathlib
import re
import sys
import unicodedata
import urllib.parse
import urllib.request
from datetime import date

RE_NGAY_DU = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def ngay_hop_le(v, ghi_chu_phu):
    """Trả ISO date nếu đủ ngày/tháng/năm. Chỉ có tháng/năm (suy từ số HĐ) -> None,
    KHÔNG tự thêm ngày 01 giả — giữ nguyên thông tin đó trong ghi_chu thay vì bịa ngày."""
    if isinstance(v, str):
        if RE_NGAY_DU.match(v):
            return v, ghi_chu_phu
        return None, (ghi_chu_phu + f" [Ngày ký chỉ biết tháng/năm: {v}]").strip()
    if v:
        return v.isoformat(), ghi_chu_phu
    return None, ghi_chu_phu

import openpyxl

from migrate.quet_hop_dong import POE, doc_file, khong_dau

ROOT = pathlib.Path(__file__).resolve().parent.parent
EXCEL = ROOT / "GWT_goi_bao_tri_tu_hop_dong_2026-07-17 (1).xlsx"

GWT_LOAI = {"0339946388", "0110530659"}
RE_A = re.compile(r"BÊN\s+(?:MUA|BÁN)\s*\(\s*A\s*\)", re.I)
RE_B = re.compile(r"BÊN\s+(?:MUA|BÁN)\s*\(\s*B\s*\)", re.I)
RE_SDT = re.compile(r"(?:SĐT|Số điện thoại|Điện thoại)\s*:?\s*(0\d{9,10})", re.I)


def nfc(s):
    return unicodedata.normalize("NFC", str(s or ""))


def trich_sdt(d):
    """d = thư mục khách -> SĐT khớp khối (A), hoặc None."""
    for f in sorted(d.rglob("*.docx"), key=lambda p: 0 if "hop dong" in khong_dau(p.name) or "hdmb" in khong_dau(p.name) else 1):
        if f.name.startswith("~$"):
            continue
        txt = doc_file(f)
        ma = RE_A.search(txt)
        if not ma:
            continue
        mb = RE_B.search(txt, ma.end())
        khoi = txt[ma.end(): mb.start() if mb else ma.end() + 400]
        m = RE_SDT.search(khoi)
        if m and m.group(1) not in GWT_LOAI:
            return m.group(1)
    return None


def sb():
    env = dict(l.strip().split("=", 1) for l in (ROOT / "app-cskh/.env.local").read_text().splitlines()
               if l.strip() and not l.startswith("#") and "=" in l)
    return env["NEXT_PUBLIC_SUPABASE_URL"], env["SUPABASE_SERVICE_ROLE_KEY"]


def get(url, key, path):
    path = urllib.parse.quote(path, safe='?&=,.*"')
    r = urllib.request.Request(f"{url}/rest/v1/{path}", headers={"apikey": key, "Authorization": f"Bearer {key}"})
    return json.load(urllib.request.urlopen(r))


def post(url, key, table, rows):
    r = urllib.request.Request(
        f"{url}/rest/v1/{table}",
        data=json.dumps(rows).encode(), method="POST",
        headers={"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json",
                 "Prefer": "return=minimal"})
    urllib.request.urlopen(r)


LOAI_MAP = {"HỢP ĐỒNG": "hop_dong", "TẶNG/NỘI BỘ": "tang_noi_bo"}


def main():
    ghi = "--ghi" in sys.argv
    url, key = sb()

    customers = get(url, key, "customers?select=id,primary_phone")
    by_phone = {c["primary_phone"]: c["id"] for c in customers if c["primary_phone"]}
    poe_serials = {r["customer_id"]: r["internal_code"] for r in
                   get(url, key, "installed_base?select=customer_id,internal_code&internal_code=in.(WH15A,WH30A)")}

    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = [nfc(c.value) for c in ws[1]]
    I = {h: i for i, h in enumerate(hdr)}
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]

    khach_dirs = {nfc(q.name): q for p in POE.iterdir() if p.is_dir() for q in p.iterdir() if q.is_dir()}

    them, thong_ke = [], {"khop_du": 0, "khop_thieu_may": 0, "khach_moi": 0, "khong_ro_sdt": 0}
    for r in rows:
        ten_folder = nfc(r[0])
        loai = LOAI_MAP.get(nfc(r[I["LOẠI GÓI"]]), "tang_noi_bo")
        d = khach_dirs.get(ten_folder)
        sdt = trich_sdt(d) if d else None

        cid = by_phone.get(sdt) if sdt else None
        if cid and cid in poe_serials:
            thong_ke["khop_du"] += 1
        elif cid:
            thong_ke["khop_thieu_may"] += 1
        elif sdt:
            thong_ke["khach_moi"] += 1
        else:
            thong_ke["khong_ro_sdt"] += 1

        ngay_ky, ghi_chu = ngay_hop_le(r[I["NGÀY KÝ HĐ"]], nfc(r[I["ĐỐI CHIẾU (ghi chú của bạn)"]]))
        them.append({
            "customer_id": cid,
            "source_folder": ten_folder,
            "source_customer_name": None if cid else ten_folder.split(". ", 1)[-1],
            "source_phone": sdt,
            "bo_may": nfc(r[I["Bộ"]]) or None,
            "loai_goi": loai,
            "ngay_ky_hd": ngay_ky,
            "so_nam": r[I["Số năm (HĐ)"]],
            "chu_ky_thang": r[I["Chu kỳ tháng (HĐ)"]],
            "tong_lan": int(r[I["→ TỔNG LẦN (tính)"]]) if r[I["→ TỔNG LẦN (tính)"]] else None,
            "ghi_chu": ghi_chu or None,
        })

    print(f"═══ {len(them)} gói bảo trì sẽ nạp ═══")
    print(f"  Khớp khách + đã có máy POE : {thong_ke['khop_du']}")
    print(f"  Khớp khách, thiếu máy POE  : {thong_ke['khop_thieu_may']}")
    print(f"  Khách hoàn toàn mới (có SĐT, không có trong DB) : {thong_ke['khach_moi']}")
    print(f"  Không trích được SĐT       : {thong_ke['khong_ro_sdt']}")
    tang = sum(1 for t in them if t["loai_goi"] == "tang_noi_bo")
    print(f"  Loại gói: hợp đồng={len(them)-tang} · tặng/nội bộ={tang}")

    if not ghi:
        print("\n(dry-run — thêm --ghi để ghi thật)")
        return
    for i in range(0, len(them), 50):
        post(url, key, "maintenance_plan", them[i:i + 50])
    print(f"\nĐÃ GHI {len(them)} gói vào maintenance_plan.")


if __name__ == "__main__":
    main()
