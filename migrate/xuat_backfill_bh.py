"""Xuất Excel đối chiếu BACKFILL bảo hành combo: MẸ (cũ, đã gỡ) → CON (mới, đã kích hoạt).

Cho user soi đúng thay đổi ở migration 24: trước đây BH nằm ở mã bộ MẸ, nay chuyển
xuống từng thiết bị CON. Đọc:
  - warranty_bak_combo_backfill_20260811 (bản backup TRƯỚC backfill) -> BH mẹ cũ
  - warranty hiện tại -> BH con mới
  - installed_base -> cấu trúc mẹ/con · cs_customers -> tên khách

Mỗi dòng 1 thiết bị con: mã bộ mẹ + BH mẹ CŨ (đã gỡ) + serial con + BH con MỚI + con
đó là 'mới kích hoạt' hay 'đã có sẵn'.

Chạy:  .venv/bin/python -m migrate.xuat_backfill_bh [--out <xlsx>]
"""

import pathlib
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from migrate.doi_chieu_khach_sales import sb, doc_het

ROOT = pathlib.Path(__file__).resolve().parent.parent
HEAD = PatternFill("solid", fgColor="1F4E79")
XANH = PatternFill("solid", fgColor="E2EFDA")   # con mới kích hoạt


def bang(ws, headers, widths, rows, to_mau=None):
    ws.append(headers)
    for j in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=j)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for r in rows:
        ws.append(r)
        fill = to_mau(r) if to_mau else None
        if fill:
            for j in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=j).fill = fill
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def main():
    out = ROOT / "migrate/_out/backfill_bh_combo.xlsx"
    if "--out" in sys.argv:
        out = pathlib.Path(sys.argv[sys.argv.index("--out") + 1])
    out.parent.mkdir(parents=True, exist_ok=True)

    url, key = sb()
    ib = doc_het(url, key, "installed_base?select=serial,internal_code,parent_serial,customer_id,install_date")
    bak = {w["serial"]: w for w in doc_het(url, key,
           "warranty_bak_combo_backfill_20260811?select=serial,start_date,full_end,core_end")}
    cur = {w["serial"]: w for w in doc_het(url, key, "warranty?select=serial,start_date,full_end,core_end")}
    kh = {c["id"]: c for c in doc_het(url, key, "cs_customers?select=id,full_name,primary_phone")}

    con_theo_me = {}
    for r in ib:
        if r.get("parent_serial"):
            con_theo_me.setdefault(r["parent_serial"], []).append(r)

    rows = []
    for me_serial, cons in sorted(con_theo_me.items()):
        me = next((r for r in ib if r["serial"] == me_serial), None)
        if not me or me.get("internal_code") not in ("WH15A", "WH30A"):
            continue
        bh_me = bak.get(me_serial) or {}
        k = kh.get(me.get("customer_id")) or {}
        for c in sorted(cons, key=lambda x: x.get("internal_code") or ""):
            bh_con = cur.get(c["serial"]) or {}
            da_co_truoc = c["serial"] in bak
            rows.append([
                me_serial, me.get("internal_code"), k.get("full_name"), k.get("primary_phone"),
                bh_me.get("start_date"), bh_me.get("full_end"), bh_me.get("core_end"),
                c.get("internal_code"), c["serial"],
                bh_con.get("start_date"), bh_con.get("full_end"), bh_con.get("core_end"),
                "đã có sẵn" if da_co_truoc else "mới kích hoạt",
            ])

    wb = Workbook(); wb.remove(wb.active)
    bang(wb.create_sheet("Backfill BH mẹ→con"),
         ["Mã bộ (MẸ)", "Combo", "Khách", "SĐT",
          "MẸ cũ: bắt đầu", "MẸ cũ: hết máy", "MẸ cũ: hết lõi",
          "Thiết bị (mã)", "Serial CON",
          "CON mới: bắt đầu", "CON mới: hết máy", "CON mới: hết lõi", "Con"],
         [20, 8, 24, 13, 13, 13, 13, 15, 22, 13, 13, 13, 14], rows,
         to_mau=lambda r: XANH if r[12] == "mới kích hoạt" else None)
    wb.save(out)

    n_bo = len({r[0] for r in rows})
    print(f"Bộ: {n_bo} · dòng con: {len(rows)} · mới kích hoạt: {sum(1 for r in rows if r[12]=='mới kích hoạt')} · đã có sẵn: {sum(1 for r in rows if r[12]=='đã có sẵn')}")
    print(f"-> {out}")


if __name__ == "__main__":
    main()
