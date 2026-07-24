"""Cập nhật DB project MỚI (GWT-SalesTracking) từ các file user đã sửa tay 2026-07-24.

Nguồn (user chốt: file sửa tay là sự thật; ca xung đột -> HỎI user, không tự chọn):
- GWT_soat_3_bang_2026-07-16.xlsx: tab KHÁCH HÀNG + "Khách cần bổ sung" (địa chỉ/SĐT/tỉnh
  — cả cột gốc lẫn cột "→"), SERIAL-MÁY (ngày lắp), TICKET (serial + trạng thái).
- GWT_22_serial_can_gan_khach_Odoo_2026-07-15.xlsx: gán khách + SĐT cho 22 serial tồn kho.

Target: project mới (DEST_URL/KEY từ migrate/.env.migrate). SO với DB, chỉ đổi chỗ khác.
Mỗi phần in 2 nhóm: ÁP ĐƯỢC (rõ ràng) + XUNG ĐỘT (cần user quyết).

Chạy:  .venv/bin/python -m migrate.cap_nhat_2026_07_24 <phan>          # dry-run
       .venv/bin/python -m migrate.cap_nhat_2026_07_24 <phan> --ghi    # ghi
  <phan> = khach | ticket | serial | gankhach
"""

import json
import pathlib
import re
import sys
import unicodedata
import urllib.parse
import urllib.request

import openpyxl

from migrate.nap_excel_sua import normalize_phone, parse_date_cell

ROOT = pathlib.Path(__file__).resolve().parent.parent
SOAT = ROOT / "GWT_soat_3_bang_2026-07-16.xlsx"
GAN = ROOT / "GWT_22_serial_can_gan_khach_Odoo_2026-07-15.xlsx"


# User chốt 2026-07-24: F00000365V9l00010139 (l thường) là GÕ NHẦM của
# F00000365V9I00010139 (I hoa) — 1 máy. Bỏ serial l thường, mọi tham chiếu -> I hoa.
SERIAL_FIX = {"F00000365V9l00010139": "F00000365V9I00010139"}


def nfc(s):
    return unicodedata.normalize("NFC", str(s or ""))


def has(v):
    return v is not None and str(v).strip() != ""


def dest():
    envf = ROOT / "migrate/.env.migrate"
    d = {}
    for line in envf.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1)
            d[k.strip()] = v.strip()
    url, key = d.get("DEST_URL"), d.get("DEST_SERVICE_KEY")
    if not url or not key:
        sys.exit("❌ Thiếu DEST trong migrate/.env.migrate")
    if "qynpywysgltspmgnhhga" in url:
        sys.exit("❌ DEST là Masterdata — dừng (target phải project mới).")
    return url, key


def rest(url, key, path, method="GET", body=None, prefer=None):
    p = urllib.parse.quote(path, safe='?&=,.*"')
    headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}
    if prefer:
        headers["Prefer"] = prefer
    req = urllib.request.Request(f"{url}/rest/v1/{p}", method=method,
                                 data=json.dumps(body).encode() if body is not None else None,
                                 headers=headers)
    with urllib.request.urlopen(req) as r:
        raw = r.read()
        return json.loads(raw) if raw else None


def doc_het(url, key, res):
    rows, off = [], 0
    while True:
        sep = "&" if "?" in res else "?"
        batch = rest(url, key, f"{res}{sep}limit=1000&offset={off}")
        rows += batch
        if len(batch) < 1000:
            return rows
        off += 1000


# ═══ PHẦN KHÁCH ═══════════════════════════════════════════════════════════════
def phan_khach(url, key, ghi):
    db = {c["id"]: c for c in doc_het(url, key, "cs_customers?select=id,full_name,primary_phone,address,province,notes")}
    phone_owner = {c["primary_phone"]: c["id"] for c in db.values() if c["primary_phone"]}

    wb = openpyxl.load_workbook(SOAT, data_only=True)
    ap, xung_dot = [], []
    for sheet in ("KHÁCH HÀNG", "Khách cần bổ sung dữ liệu"):
        for r in wb[sheet].iter_rows(min_row=2, values_only=True):
            cid = str(r[0] or "").strip()
            if not cid or cid not in db:
                continue
            c = db[cid]
            patch, notes_ap = {}, []

            # SĐT: ưu tiên →[12] rồi gốc[2]
            raw_sdt = r[12] if has(r[12]) else (r[2] if has(r[2]) else None)
            if raw_sdt is not None:
                p, err = normalize_phone(raw_sdt)
                if err:
                    xung_dot.append((c["full_name"], f"SĐT không đọc được: {err}"))
                elif p != c["primary_phone"]:
                    other = phone_owner.get(p)
                    if other and other != cid:
                        xung_dot.append((c["full_name"], f"SĐT {p} đã là của khách khác (id {other[:8]}) — cần gộp/kiểm"))
                    elif c["primary_phone"] and c["primary_phone"] != p:
                        xung_dot.append((c["full_name"], f"SĐT DB {c['primary_phone']} ≠ file {p} — cả 2 có giá trị, chọn?"))
                    else:  # DB rỗng, file có -> bổ sung
                        patch["primary_phone"] = p
                        patch["needs_phone"] = False
                        notes_ap.append(f"SĐT +{p}")

            # Địa chỉ: →[14] rồi gốc[3]. File có + khác DB -> áp (user sửa = đúng). File rỗng -> giữ DB.
            dc = r[14] if has(r[14]) else (r[3] if has(r[3]) else None)
            if dc is not None and nfc(dc).strip() != nfc(c["address"]).strip():
                patch["address"] = nfc(dc).strip()
                notes_ap.append("địa chỉ")

            # Tỉnh: →[15] rồi gốc[4]
            tinh = r[15] if has(r[15]) else (r[4] if has(r[4]) else None)
            if tinh is not None and nfc(tinh).strip() != nfc(c["province"]).strip():
                patch["province"] = nfc(tinh).strip()
                notes_ap.append("tỉnh")

            # Ghi chú của bạn -> notes. BỎ marker review ("OK"/"đã sửa"...) — đó là dấu
            # user tick đã kiểm, không phải nội dung notes của khách.
            gc = nfc(r[16]).strip()
            MARKER = {"ok", "đã sửa", "da sua", "done", "x", "đã ok", "oke", "okie"}
            if has(r[16]) and gc.lower() not in MARKER and gc != nfc(c["notes"]).strip():
                patch["notes"] = gc
                notes_ap.append("ghi chú")

            if patch:
                ap.append((cid, c["full_name"], "; ".join(notes_ap), patch))

    print(f"═══ KHÁCH — ÁP ĐƯỢC: {len(ap)} · XUNG ĐỘT: {len(xung_dot)} ═══")
    for _, ten, mota, _ in ap[:30]:
        print(f"  ✎ {nfc(ten)[:26]:28} {mota}")
    if len(ap) > 30:
        print(f"  … còn {len(ap)-30}")
    if xung_dot:
        print(f"\n  ⚠️ XUNG ĐỘT (cần bạn quyết):")
        for ten, ly in xung_dot:
            print(f"    {nfc(ten)[:26]:28} {ly}")

    if ghi:
        for cid, _, _, patch in ap:
            rest(url, key, f"cs_customers?id=eq.{cid}", method="PATCH", body=patch, prefer="return=minimal")
        print(f"\n✓ Đã cập nhật {len(ap)} khách.")
    else:
        print("\n(dry-run — thêm --ghi để ghi)")


# ═══ PHẦN TICKET ══════════════════════════════════════════════════════════════
def chuan_state(v):
    s = nfc(v).strip().capitalize()
    return s if s in ("Open", "Done", "Cancel") else None


def phan_ticket(url, key, ghi):
    db = {t["ticket_code"]: t for t in doc_het(url, key, "tickets?select=ticket_code,state,serial,source_serial")}
    ib_serials = {r["serial"] for r in doc_het(url, key, "installed_base?select=serial")}
    wb = openpyxl.load_workbook(SOAT, data_only=True)
    ap, xung_dot = [], []
    for r in wb["TICKET"].iter_rows(min_row=2, values_only=True):
        tc = str(r[0] or "").strip()
        if not tc or tc not in db:
            continue
        t = db[tc]
        patch, notes_ap = {}, []
        # trạng thái →[14]
        if has(r[14]):
            st = chuan_state(r[14])
            if not st:
                xung_dot.append((tc, f"trạng thái không hợp lệ: {r[14]!r}"))
            elif st != t["state"]:
                if t["state"] == "Cancel" and st == "Done":
                    xung_dot.append((tc, f"DB đang Cancel, file muốn Done — chắc chắn?"))
                else:
                    patch["state"] = st
                    notes_ap.append(f"state {t['state']}→{st}")
        # serial →[13]
        if has(r[13]):
            s = SERIAL_FIX.get(nfc(r[13]).strip(), nfc(r[13]).strip())
            if s not in ib_serials:
                xung_dot.append((tc, f"serial {s} không có trong installed_base"))
            elif s != t["serial"]:
                patch["serial"] = s
                notes_ap.append(f"serial→{s}")
        if patch:
            ap.append((tc, "; ".join(notes_ap), patch))

    print(f"═══ TICKET — ÁP ĐƯỢC: {len(ap)} · XUNG ĐỘT: {len(xung_dot)} ═══")
    for tc, mota, _ in ap:
        print(f"  ✎ {tc:14} {mota}")
    if xung_dot:
        print(f"\n  ⚠️ XUNG ĐỘT:")
        for tc, ly in xung_dot:
            print(f"    {tc:14} {ly}")
    if ghi:
        for tc, _, patch in ap:
            rest(url, key, f"tickets?ticket_code=eq.{tc}", method="PATCH", body=patch, prefer="return=minimal")
        print(f"\n✓ Đã cập nhật {len(ap)} ticket.")
    else:
        print("\n(dry-run — thêm --ghi để ghi)")


# ═══ PHẦN GÁN KHÁCH (GWT_22: serial tồn kho có ticket -> thêm ib + gán khách) ═══
def phan_gankhach(url, key, ghi):
    ib = {r["serial"]: r for r in doc_het(url, key, "installed_base?select=serial,customer_id")}
    cust = doc_het(url, key, "cs_customers?select=id,full_name,primary_phone")
    by_phone = {c["primary_phone"]: c["id"] for c in cust if c["primary_phone"]}
    catalog = {r["Mã nội bộ"] for r in doc_het(url, key, 'catalog_item?select="Mã nội bộ"')}

    wb = openpyxl.load_workbook(GAN, data_only=True)
    rows = [r for r in wb["Sửa Odoo"].iter_rows(values_only=True) if r[0] and str(r[0]).strip().isdigit()]

    them_ib, tao_khach, xung_dot, bo_qua = [], {}, [], 0
    for r in rows:
        serial = nfc(r[2]).strip()
        if serial in SERIAL_FIX:   # V9l gõ nhầm -> bỏ (V9I đã có dòng riêng)
            continue
        ten = nfc(r[3]).strip()
        sdt, _ = normalize_phone(r[4]) if has(r[4]) else (None, None)
        m = re.search(r"\[([^\]]+)\]", nfc(r[5]))
        icode = m.group(1) if m else None

        # tìm/tạo khách theo SĐT (KHÔNG khớp tên — bài học 2 khách tên Yến)
        cid = by_phone.get(sdt) if sdt else None
        if serial in ib:
            # đã có máy: chỉ xét gán khách nếu đang thiếu
            if ib[serial]["customer_id"]:
                bo_qua += 1
            elif cid:
                them_ib.append(("update", serial, {"customer_id": cid}, f"gán khách có sẵn (SĐT {sdt})"))
            elif sdt:
                tao_khach.setdefault(sdt, (ten, serial))
            else:
                xung_dot.append((serial, f"máy đã có, thiếu khách, GWT_22 không có SĐT ({ten}) — gán tay"))
            continue
        # máy CHƯA có -> thêm ib
        if not icode:
            xung_dot.append((serial, f"không đọc được internal_code từ SP: {nfc(r[5])[:30]}"))
            continue
        if not sdt:
            xung_dot.append((serial, f"máy mới nhưng khách '{ten}' không có SĐT — cần gán tay"))
            continue
        if not cid:
            tao_khach.setdefault(sdt, (ten, serial))
        them_ib.append(("insert", serial, {
            "serial": serial, "internal_code": icode if icode in catalog else None,
            "source_product_code": icode,
            "model_freetext": None if icode in catalog else nfc(r[5]),
            "_sdt": sdt,  # resolve customer_id sau khi tạo khách
        }, f"thêm máy {icode} + gán khách (SĐT {sdt})"))

    print(f"═══ GÁN KHÁCH — thêm máy: {sum(1 for x in them_ib if x[0]=='insert')} · "
          f"gán khách máy có sẵn: {sum(1 for x in them_ib if x[0]=='update')} · "
          f"tạo khách mới: {len(tao_khach)} · bỏ qua (đã đủ): {bo_qua} · xung đột: {len(xung_dot)} ═══\n")
    for typ, s, _, mota in them_ib:
        print(f"  {'＋máy' if typ=='insert' else '✎khách'} {s:24} {mota}")
    if tao_khach:
        print(f"\n  Khách MỚI sẽ tạo ({len(tao_khach)}):")
        for sdt, (ten, s) in tao_khach.items():
            print(f"    {nfc(ten)[:26]:28} SĐT {sdt}")
    if xung_dot:
        print(f"\n  ⚠️ XUNG ĐỘT (gán tay):")
        for s, ly in xung_dot:
            print(f"    {s:24} {ly}")

    if ghi:
        # 1. tạo khách mới
        new_ids = {}
        for sdt, (ten, s) in tao_khach.items():
            got = rest(url, key, "cs_customers", method="POST",
                       body=[{"full_name": ten, "primary_phone": sdt, "source": "GWT_22 ticket 2026-07-24"}],
                       prefer="return=representation")
            new_ids[sdt] = got[0]["id"]
        by_phone.update(new_ids)
        # 2. thêm máy + gán
        for typ, s, patch, _ in them_ib:
            if typ == "insert":
                sdt = patch.pop("_sdt")
                patch["customer_id"] = by_phone.get(sdt)
                rest(url, key, "installed_base", method="POST", body=[patch], prefer="return=minimal,resolution=merge-duplicates")
            else:
                rest(url, key, f"installed_base?serial=eq.{urllib.parse.quote(s)}", method="PATCH", body=patch, prefer="return=minimal")
        print(f"\n✓ Tạo {len(tao_khach)} khách + thêm/gán {len(them_ib)} máy.")
    else:
        print("\n(dry-run — thêm --ghi để ghi)")


def main():
    if len(sys.argv) < 2 or sys.argv[1] not in ("khach", "ticket", "serial", "gankhach"):
        sys.exit("Dùng: cap_nhat_2026_07_24 <khach|ticket|serial|gankhach> [--ghi]")
    phan = sys.argv[1]
    ghi = "--ghi" in sys.argv
    url, key = dest()
    print(f"DEST: {url}\n{'='*60}")
    {"khach": phan_khach, "ticket": phan_ticket, "gankhach": phan_gankhach}.get(phan, lambda *a: print("phần chưa làm"))(url, key, ghi)


if __name__ == "__main__":
    main()
