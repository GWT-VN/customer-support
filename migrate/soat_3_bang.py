"""Xuất Excel soát toàn bộ 3 bảng: Khách hàng · Serial-Máy · Ticket.

Mỗi dòng được đánh dấu OK / CẦN SỬA + lý do, kèm cột "SỬA THÀNH" để người
duyệt điền trực tiếp rồi gửi lại file — script nạp ngược sẽ đọc theo khoá
(id / serial / ticket_code) ở cột A.

Chạy:  .venv/bin/python -m migrate.soat_3_bang

⚠️ File xuất ra CHỨA TÊN/SĐT/ĐỊA CHỈ KHÁCH -> đã gitignore, KHÔNG commit.
"""

import json
import pathlib
import urllib.parse
import urllib.request
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from migrate import quality

ROOT = pathlib.Path(__file__).resolve().parent.parent
TODAY = date(2026, 7, 16)
OUT = ROOT / f"GWT_soat_3_bang_{TODAY.isoformat()}.xlsx"


# ── Kết nối ──────────────────────────────────────────────────────────────────
def load_env():
    env = {}
    for line in (ROOT / "app-cskh/.env.local").read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip()
    return env


def fetch_all(url, key, resource, order):
    """Đọc HẾT bảng (phân trang 1000 dòng/lượt cho chắc)."""
    rows, offset = [], 0
    while True:
        sep = "&" if "?" in resource else "?"
        q = f"{url}/rest/v1/{resource}{sep}order={urllib.parse.quote(order)}&limit=1000&offset={offset}"
        req = urllib.request.Request(q, headers={"apikey": key, "Authorization": f"Bearer {key}"})
        batch = json.load(urllib.request.urlopen(req))
        rows += batch
        if len(batch) < 1000:
            return rows
        offset += 1000


# ── Soát từng domain ─────────────────────────────────────────────────────────
def audit_khach(c, so_may):
    """-> (tình trạng, lý do, ưu tiên). Tái dùng luật của quality.py."""
    row = dict(c)
    row["so_may"] = so_may
    note, muc = quality.audit(row)
    return ("CẦN SỬA" if note else "OK"), note, muc


def audit_may(m):
    notes, worst = [], 0  # 0 ok · 1 nhẹ · 2 vừa · 3 cao
    la_con = bool(m.get("parent_serial"))   # serial con trong bộ lọc tổng
    if not m.get("customer_id"):
        notes.append("KHÔNG GẮN KHÁCH — máy lắp ngoài thị trường phải thuộc về một khách.")
        worst = 3
    if not m.get("internal_code"):
        if m.get("source_product_code"):
            notes.append(f"Mã Odoo “{m['source_product_code']}” KHÔNG tra được ra mã nội bộ "
                         "— nếu là mã hãng/kho thì bổ sung vào supplier_code, nếu máy ngoài catalog thì ghi rõ.")
            worst = max(worst, 2)
        else:
            notes.append("Máy ngoài catalog (chỉ có tên tự do) — nếu thực ra có mã nội bộ thì điền.")
            worst = max(worst, 1)
    if not m.get("install_date"):
        if la_con:
            notes.append("Thiếu ngày lắp (kế thừa mẹ — mẹ cũng thiếu, sửa ở dòng serial MẸ).")
            worst = max(worst, 1)
        else:
            notes.append("THIẾU NGÀY LẮP — đây là mốc tính lịch thay lõi khi chưa có log thay.")
            worst = max(worst, 2)
    if not m.get("warranty_activated"):
        if la_con:
            notes.append("BH tính theo bộ nhưng serial MẸ chưa kích hoạt — xử lý ở dòng serial mẹ.")
            worst = max(worst, 1)
        else:
            notes.append("Chưa kích hoạt bảo hành — xác nhận có chủ đích không (máy cũ/hết hạn?).")
            worst = max(worst, 2)
    elif m.get("warranty_full_end") is None:
        if m.get("co_chinh_sach_bh") is False:
            notes.append("Đã kích hoạt BH nhưng SP KHÔNG có chính sách số năm (master data ghi "
                         "“Không áp dụng” hoặc chưa khai) — xác nhận với nghiệp vụ.")
        else:
            notes.append("Đã kích hoạt BH nhưng không tính được ngày hết hạn — thiếu số năm trong product_warranty.")
        worst = max(worst, 2)
    if m.get("status") and m["status"] != "active":
        notes.append(f"Trạng thái máy = {m['status']} — kiểm tra còn theo dõi không.")
        worst = max(worst, 1)
    if m.get("needs_phone"):
        notes.append("(Khách của máy này đang thiếu/sai SĐT — sửa ở sheet KHÁCH HÀNG.)")
        worst = max(worst, 1)
    muc = {0: "", 1: "Thấp", 2: "Vừa", 3: "CAO"}[worst]
    return ("CẦN SỬA" if notes else "OK"), " ".join(notes), muc


def audit_ticket(t):
    notes, worst = [], 0
    if t.get("may_khong_trong_he_thong"):
        notes.append(f"Serial Odoo “{t.get('source_serial')}” KHÔNG có trong hệ máy đã lắp "
                     "(Odoo để máy ở tồn kho dù đang nhà khách) — gắn serial đúng hoặc bổ sung máy.")
        worst = 3
    elif not t.get("serial") and not t.get("source_serial"):
        notes.append("Ticket không gắn máy nào — nếu biết serial thì điền.")
        worst = max(worst, 2)
    if not t.get("customer_id"):
        notes.append("Không gắn khách trong hệ thống"
                     + (f" (Odoo ghi: “{t['source_customer']}”)" if t.get("source_customer") else "")
                     + " — tìm/tạo khách rồi gắn.")
        worst = max(worst, 2)
    if t.get("state") == "Open":
        created = t.get("created_at")
        if created:
            d = datetime.fromisoformat(created.replace("Z", "+00:00")).date()
            days = (TODAY - d).days
            if days > 30:
                notes.append(f"Ticket MỞ đã {days} ngày — còn xử lý thật không hay quên đóng?")
                worst = max(worst, 2)
    if not (t.get("description") or "").strip():
        notes.append("Thiếu mô tả lỗi — không biết khách báo gì.")
        worst = max(worst, 2)
    muc = {0: "", 1: "Thấp", 2: "Vừa", 3: "CAO"}[worst]
    return ("CẦN SỬA" if notes else "OK"), " ".join(notes), muc


# ── Excel ────────────────────────────────────────────────────────────────────
HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
EDIT_FILL = PatternFill("solid", fgColor="DDEBF7")   # cột người duyệt điền
CAO_FILL = PatternFill("solid", fgColor="FFC7CE")
VUA_FILL = PatternFill("solid", fgColor="FFEB9C")
THAP_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FONT = Font(color="006100")
BAD_FONT = Font(color="9C0006", bold=True)
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def write_sheet(wb, title, headers, widths, rows, n_edit_cols):
    """rows = list[list]; 3 cột cuối trước nhóm edit là TÌNH TRẠNG/LÝ DO/ƯU TIÊN."""
    ws = wb.create_sheet(title)
    ws.append(headers)
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if j > len(headers) - n_edit_cols:
            cell.fill = PatternFill("solid", fgColor="2E75B6")
    idx_status = headers.index("TÌNH TRẠNG") + 1
    idx_reason = headers.index("LÝ DO / CẦN SỬA GÌ") + 1
    idx_muc = headers.index("ƯU TIÊN") + 1
    for r in rows:
        ws.append(r)
        i = ws.max_row
        muc = ws.cell(row=i, column=idx_muc).value
        fill = {"CAO": CAO_FILL, "Vừa": VUA_FILL, "Thấp": THAP_FILL}.get(muc)
        for j in range(1, len(headers) + 1):
            c = ws.cell(row=i, column=j)
            c.border = THIN
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if fill and j in (idx_status, idx_reason, idx_muc):
                c.fill = fill
            if j > len(headers) - n_edit_cols:
                c.fill = EDIT_FILL if not fill or j > idx_muc else c.fill
        st = ws.cell(row=i, column=idx_status)
        st.font = BAD_FONT if st.value == "CẦN SỬA" else OK_FONT
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    return ws


def main():
    env = load_env()
    url, key = env["NEXT_PUBLIC_SUPABASE_URL"], env["SUPABASE_SERVICE_ROLE_KEY"]

    customers = fetch_all(url, key, "customers", "full_name")
    machines = fetch_all(url, key, "v_installed_base", "customer_name,serial")
    tickets = fetch_all(url, key, "v_tickets", "ticket_code")
    contacts = fetch_all(url, key, "customer_contacts", "customer_id")

    may_cua = {}
    for m in machines:
        if m.get("customer_id"):
            may_cua[m["customer_id"]] = may_cua.get(m["customer_id"], 0) + 1
    sdt_phu = {}
    for ct in contacts:
        sdt_phu[ct["customer_id"]] = sdt_phu.get(ct["customer_id"], 0) + 1

    wb = Workbook()
    wb.remove(wb.active)

    # ── TỔNG QUAN ────────────────────────────────────────────────────────
    ov = wb.create_sheet("TỔNG QUAN")

    # ── KHÁCH HÀNG ───────────────────────────────────────────────────────
    kh_rows = []
    for c in customers:
        status, note, muc = audit_khach(c, may_cua.get(c["id"], 0))
        kh_rows.append([
            c["id"], c["full_name"], c.get("primary_phone"), c.get("address"),
            c.get("province"), c.get("source"), may_cua.get(c["id"], 0),
            sdt_phu.get(c["id"], 0), c.get("notes"),
            status, note, muc,
            None, None, None, None, None,
        ])
    kh_rows.sort(key=lambda r: ({"CAO": 0, "Vừa": 1, "Thấp": 2, "": 3}[r[11]], r[1] or ""))
    write_sheet(
        wb, "KHÁCH HÀNG",
        ["id (KHÔNG SỬA)", "Tên khách", "SĐT chính", "Địa chỉ", "Tỉnh/TP", "Nguồn",
         "Số máy", "SĐT phụ đã lưu", "Ghi chú hệ thống",
         "TÌNH TRẠNG", "LÝ DO / CẦN SỬA GÌ", "ƯU TIÊN",
         "→ SĐT sửa thành", "→ Tên sửa thành", "→ Địa chỉ sửa thành",
         "→ Tỉnh/TP sửa thành", "→ Ghi chú của bạn"],
        [10, 24, 14, 40, 12, 10, 7, 8, 20, 11, 55, 8, 14, 20, 40, 12, 30],
        kh_rows, 5)

    # ── SERIAL-MÁY ───────────────────────────────────────────────────────
    may_rows = []
    for m in machines:
        status, note, muc = audit_may(m)
        if m.get("warranty_activated"):
            han = m.get("warranty_full_end") or "không rõ hạn"
            bh = f"✔ theo bộ mẹ ({han})" if m.get("bh_theo_me") else f"✔ {han}"
        else:
            bh = "CHƯA"
        may_rows.append([
            m["serial"], m.get("internal_code"), m.get("product_name"),
            m.get("source_product_code"), m.get("customer_name"), m.get("primary_phone"),
            m.get("install_date"), m.get("install_address"), m.get("parent_serial"),
            m.get("status"), bh, m.get("warranty_core_end"),
            status, note, muc,
            None, None, None, None,
        ])
    may_rows.sort(key=lambda r: ({"CAO": 0, "Vừa": 1, "Thấp": 2, "": 3}[r[14]], r[4] or "", r[0]))
    write_sheet(
        wb, "SERIAL-MÁY",
        ["Serial (KHÔNG SỬA)", "Mã nội bộ", "Tên máy", "Mã gốc Odoo", "Khách", "SĐT khách",
         "Ngày lắp", "Đ/c lắp đặt", "Serial cha", "Trạng thái", "BH kích hoạt (hết hạn máy)",
         "Hết hạn lõi",
         "TÌNH TRẠNG", "LÝ DO / CẦN SỬA GÌ", "ƯU TIÊN",
         "→ Mã nội bộ đúng", "→ Ngày lắp đúng (YYYY-MM-DD)", "→ Kích hoạt BH? (x = có)",
         "→ Ghi chú của bạn"],
        [22, 16, 28, 14, 22, 13, 11, 30, 20, 9, 15, 11, 11, 55, 8, 14, 14, 10, 30],
        may_rows, 4)

    # ── TICKET ───────────────────────────────────────────────────────────
    tk_rows = []
    for t in tickets:
        status, note, muc = audit_ticket(t)
        tk_rows.append([
            t["ticket_code"],
            (t.get("created_at") or "")[:10], t.get("state"), t.get("ticket_type"),
            t.get("serial") or t.get("source_serial"), t.get("product_name"),
            t.get("customer_name"), t.get("primary_phone"),
            t.get("description"), t.get("last_note"),
            status, note, muc,
            None, None, None,
        ])
    tk_rows.sort(key=lambda r: ({"CAO": 0, "Vừa": 1, "Thấp": 2, "": 3}[r[12]], r[0]))
    write_sheet(
        wb, "TICKET",
        ["Mã ticket (KHÔNG SỬA)", "Ngày tạo", "Trạng thái", "Loại", "Serial", "Máy",
         "Khách", "SĐT khách", "Mô tả lỗi", "Ghi chú xử lý",
         "TÌNH TRẠNG", "LÝ DO / CẦN SỬA GÌ", "ƯU TIÊN",
         "→ Serial đúng", "→ Trạng thái đúng (Open/Done/Cancel)", "→ Ghi chú của bạn"],
        [14, 11, 10, 22, 22, 24, 22, 13, 45, 30, 11, 55, 8, 20, 16, 30],
        tk_rows, 3)

    # ── MÁY THIẾU LÕI (không tra được lịch thay lõi) ─────────────────────
    mf_codes = {r["internal_code"] for r in fetch_all(url, key, "v_machine_filter?select=internal_code", "internal_code")}
    thieu = {}
    for m in machines:
        code = m.get("internal_code")
        if code and code not in mf_codes:
            t = thieu.setdefault(code, {"ten": m.get("product_name"), "serials": []})
            t["serials"].append(m["serial"])
        elif not code:
            t = thieu.setdefault(f"(ngoài catalog) {m.get('product_name')}", {"ten": m.get("product_name"), "serials": []})
            t["serials"].append(m["serial"])

    tl = wb.create_sheet("MÁY THIẾU LÕI")
    tl_head = ["Mã máy (KHÔNG SỬA)", "Tên máy", "Số máy đã lắp", "Serial các máy",
               "→ Mã lõi lọc", "→ Tên lõi lọc", "→ Chu kỳ thay (tháng, vd 12 hoặc 12-24)", "→ Ghi chú"]
    tl.append(tl_head)
    for j, h in enumerate(tl_head, 1):
        c = tl.cell(row=1, column=j)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEAD_FILL if j <= 4 else PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for code, t in sorted(thieu.items()):
        first = tl.max_row + 1
        tl.append([code, t["ten"], len(t["serials"]), ", ".join(t["serials"])])
        for _ in range(7):                      # mỗi lõi 1 dòng — chừa sẵn 8 dòng/máy
            tl.append([code])
        for i in range(first, tl.max_row + 1):
            for j in range(1, len(tl_head) + 1):
                c = tl.cell(row=i, column=j)
                c.border = THIN
                c.alignment = Alignment(wrap_text=True, vertical="top")
                if j >= 5:
                    c.fill = EDIT_FILL
    for j, w in enumerate([26, 30, 8, 46, 16, 30, 18, 24], 1):
        tl.column_dimensions[get_column_letter(j)].width = w
    tl.freeze_panes = "A2"

    # ── điền TỔNG QUAN ───────────────────────────────────────────────────
    def dem(rows, i_status, i_muc):
        tong = len(rows)
        can = sum(1 for r in rows if r[i_status] == "CẦN SỬA")
        cao = sum(1 for r in rows if r[i_muc] == "CAO")
        vua = sum(1 for r in rows if r[i_muc] == "Vừa")
        return tong, tong - can, can, cao, vua

    ov["A1"] = f"SOÁT DỮ LIỆU CSKH — xuất từ Supabase GWT-Masterdata ngày {TODAY.strftime('%d/%m/%Y')}"
    ov["A1"].font = Font(bold=True, size=14)
    ov.append([])
    ov.append(["Sheet", "Tổng dòng", "OK", "CẦN SỬA", "— mức CAO", "— mức Vừa"])
    for cell in ov[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEAD_FILL
    ov.append(["KHÁCH HÀNG", *dem(kh_rows, 9, 11)[0:1], *dem(kh_rows, 9, 11)[1:]])
    ov.append(["SERIAL-MÁY", *dem(may_rows, 12, 14)[0:1], *dem(may_rows, 12, 14)[1:]])
    ov.append(["TICKET", *dem(tk_rows, 10, 12)[0:1], *dem(tk_rows, 10, 12)[1:]])
    ov.append([])
    for line in [
        "CÁCH DÙNG:",
        "1. Mỗi sheet đã lọc sẵn được (Data filter) — lọc cột TÌNH TRẠNG = CẦN SỬA, làm từ ƯU TIÊN = CAO trở xuống.",
        "2. CHỈ điền vào các cột xanh nhạt có dấu “→” (bên phải). KHÔNG sửa cột khoá đầu tiên (id/serial/mã ticket).",
        "3. Dòng OK không phải làm gì. Nếu thấy dòng OK mà thực tế sai, cứ điền cột “→” — script nạp sẽ đọc.",
        "4. Điền xong gửi lại file — tôi sẽ đọc các cột “→” và cập nhật vào hệ thống theo khoá cột A.",
        "5. Cột “→ Kích hoạt BH?”: đánh x nếu xác nhận máy đó cần kích hoạt bảo hành.",
        "6. Sheet MÁY THIẾU LÕI: các máy chưa tra được lịch thay lõi — điền MỖI LÕI 1 DÒNG "
        "(mã lõi + tên + chu kỳ tháng); đã chừa sẵn 8 dòng/máy, thiếu thì thêm dòng, thừa thì bỏ trống.",
        "7. Serial con của bộ lọc tổng (WH15A/WH30A): BH hiển thị “theo bộ mẹ” — không cần kích hoạt riêng.",
    ]:
        ov.append([line])
    ov.column_dimensions["A"].width = 110
    for w, col in zip([14, 10, 8, 10, 10, 10], "ABCDEF"):
        if col != "A":
            ov.column_dimensions[col].width = w

    wb.save(OUT)
    print(f"Đã ghi: {OUT}")
    print(f"KHÁCH HÀNG: {dem(kh_rows, 9, 11)}  (tổng, ok, cần sửa, cao, vừa)")
    print(f"SERIAL-MÁY: {dem(may_rows, 12, 14)}")
    print(f"TICKET:     {dem(tk_rows, 10, 12)}")


if __name__ == "__main__":
    main()
