"""Xuất DANH SÁCH BỘ COMBO (mẹ + con) ra Excel để user kiểm tra bộ cũ ↔ bộ mới.

KHÔNG ghi DB — chỉ đọc installed_base/warranty/cs_customers rồi xuất 1 file Excel:
  1. "Bộ (mẹ+con)"  — mỗi dòng 1 THIẾT BỊ con, gom theo bộ; cột "Kiểu" = cũ/mới
  2. "Tóm tắt"       — đếm bộ theo combo × kiểu

Kiểu: "mới" = mã bộ do hệ sinh (WH15A/WH30A + 9 số, vd WH30A202608001);
       "cũ"  = mã bundle Odoo trước đây (vd GTE30A20240002).

Chạy:  .venv/bin/python -m migrate.xuat_bo_combo
       .venv/bin/python -m migrate.xuat_bo_combo --out /đường/dẫn.xlsx
"""

import json
import pathlib
import re
import sys
import urllib.parse
import urllib.request

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent
RE_MOI = re.compile(r"^WH(15A|30A)\d{9}$")   # mã bộ hệ sinh: combo + YYYYMM + STT(3)


def sb():
    env = dict(
        l.strip().split("=", 1)
        for l in (ROOT / "app-cskh/.env.local").read_text().splitlines()
        if l.strip() and not l.startswith("#") and "=" in l
    )
    return env["NEXT_PUBLIC_SUPABASE_URL"], env["SUPABASE_SERVICE_ROLE_KEY"]


def get(url, key, path):
    path = urllib.parse.quote(path, safe='?&=,.*"<>')
    r = urllib.request.Request(f"{url}/rest/v1/{path}",
                               headers={"apikey": key, "Authorization": f"Bearer {key}"})
    return json.load(urllib.request.urlopen(r))


def doc_het(url, key, res):
    rows, off = [], 0
    while True:
        sep = "&" if "?" in res else "?"
        batch = get(url, key, f"{res}{sep}limit=1000&offset={off}")
        rows += batch
        if len(batch) < 1000:
            return rows
        off += 1000


HEAD = PatternFill("solid", fgColor="1F4E79")
XANH = PatternFill("solid", fgColor="E2EFDA")   # bộ mới
DO = PatternFill("solid", fgColor="FFC7CE")     # bộ thiếu con


def bang(ws, headers, widths, rows, to_mau=None):
    ws.append(headers)
    for j in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=j)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for r in rows:
        ws.append(r)
        fill = to_mau(r) if to_mau else None
        if fill:
            for j in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=j).fill = fill
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def main():
    out = ROOT / "migrate/_out/xuat_bo_combo.xlsx"
    if "--out" in sys.argv:
        out = pathlib.Path(sys.argv[sys.argv.index("--out") + 1])
    out.parent.mkdir(parents=True, exist_ok=True)

    url, key = sb()
    ib = doc_het(url, key, "installed_base?select=serial,internal_code,parent_serial,customer_id,install_date,status")
    wr = {w["serial"]: w for w in doc_het(url, key, "warranty?select=serial,full_end,core_end,activated")}
    kh = {c["id"]: c for c in doc_het(url, key, "cs_customers?select=id,full_name,primary_phone")}

    cha = {r["parent_serial"] for r in ib if r.get("parent_serial")}   # giá trị parent_serial = serial MẸ
    con_theo_me = {}
    for r in ib:
        if r.get("parent_serial"):
            con_theo_me.setdefault(r["parent_serial"], []).append(r)

    DU_CON = 3   # bộ combo cần đủ 3 thiết bị con (WH15A/WH30A: 3 máy · ECO: mềm + 2×UPF10)

    rows, tomtat, thieu = [], {}, []
    for me in sorted((r for r in ib if r["serial"] in cha), key=lambda x: x["serial"]):
        kieu = "mới" if RE_MOI.match(me["serial"]) else "cũ"
        combo = me.get("internal_code") or "—"
        tomtat[(combo, kieu)] = tomtat.get((combo, kieu), 0) + 1
        k = kh.get(me.get("customer_id")) or {}
        cons = sorted(con_theo_me.get(me["serial"], []), key=lambda x: x.get("internal_code") or "")
        so_con = len(cons)
        du = so_con >= DU_CON
        for c in cons:
            w = wr.get(c["serial"]) or {}
            rows.append([
                me["serial"], kieu, combo,
                k.get("full_name"), k.get("primary_phone"), me.get("install_date"),
                c.get("internal_code"), c["serial"],
                w.get("full_end"), w.get("core_end"),
                "có" if w.get("activated") else "KHÔNG",
                "" if du else f"THIẾU ({so_con}/{DU_CON})",
            ])
        if not du:
            thieu.append([
                me["serial"], combo, k.get("full_name"), k.get("primary_phone"),
                so_con, ", ".join(c.get("internal_code") or "?" for c in cons) or "—",
            ])

    # Làm mềm (GTEC-15A/30A) ĐỨNG MỘT MÌNH: chưa thuộc bộ nào -> khách mua 15A/30A hoặc
    # ECO nhưng mới kích hoạt mỗi làm mềm, thiếu các thiết bị/UPF10 còn lại.
    LAM_MEM = {'GTEC-15A01-G', 'GTEC-30A01-G'}
    mot_minh = []
    for r in ib:
        if r.get('internal_code') in LAM_MEM and not r.get('parent_serial') and r['serial'] not in cha:
            k = kh.get(r.get('customer_id')) or {}
            w = wr.get(r['serial']) or {}
            mot_minh.append([
                r['serial'], r['internal_code'], k.get('full_name'), k.get('primary_phone'),
                r.get('install_date'), 'có' if w.get('activated') else 'KHÔNG',
            ])

    wb = Workbook(); wb.remove(wb.active)
    bang(wb.create_sheet("Bộ (mẹ+con)"),
         ["Mã bộ", "Kiểu", "Combo", "Khách", "SĐT", "Ngày lắp",
          "Thiết bị (mã)", "Serial con", "BH máy hết", "BH lõi hết", "Con có BH?", "Thiếu con?"],
         [20, 7, 10, 26, 14, 12, 16, 22, 12, 12, 10, 14], rows,
         to_mau=lambda r: DO if r[11] else (XANH if r[1] == "mới" else None))
    bang(wb.create_sheet("Bộ THIẾU con"),
         ["Mã bộ", "Combo", "Khách", "SĐT", "Số con", "Thiết bị đang có (bổ sung phần thiếu)"],
         [20, 10, 26, 14, 8, 40], sorted(thieu, key=lambda r: (r[2] or "")),
         to_mau=lambda r: DO)
    bang(wb.create_sheet("Làm mềm ĐỨNG 1 MÌNH"),
         ["Serial làm mềm", "Mã", "Khách", "SĐT", "Ngày lắp", "BH kích hoạt?"],
         [22, 14, 26, 14, 12, 12], sorted(mot_minh, key=lambda r: (r[2] or "")),
         to_mau=lambda r: DO)
    tt_rows = sorted([[c, k, n] for (c, k), n in tomtat.items()], key=lambda r: (r[0], r[1]))
    bang(wb.create_sheet("Tóm tắt"), ["Combo", "Kiểu", "Số bộ"], [12, 8, 10], tt_rows)
    wb.save(out)

    n_bo = len(cha); n_moi = sum(1 for s in cha if RE_MOI.match(s))
    print(f"Tổng bộ: {n_bo} (mới: {n_moi} · cũ: {n_bo - n_moi}) · dòng thiết bị: {len(rows)} · "
          f"BỘ THIẾU con: {len(thieu)} · làm mềm đứng 1 mình: {len(mot_minh)}")
    print(f"-> {out}")


if __name__ == "__main__":
    main()
