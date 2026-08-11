"""Nạp khách công ty từ DM_KHACH_CTY: gán channel cho ca ĐÃ KHỚP + import ca CHƯA CÓ.

User chốt (2026-08-11): 5 ca khớp cs_customers -> gán channel_id luôn; 72 ca còn lại
-> import thành khách mới kèm channel + info công ty (vào notes).

Khớp: SĐT liên hệ -> primary_phone, rồi tên (nội bộ) bỏ dấu. Map "Kênh (mới)"
(l1/l2) -> dim_channel.id (❓/đa kênh -> để trống channel).

Chạy:  .venv/bin/python -m migrate.nap_khach_cty          # DRY-RUN (chỉ in)
       .venv/bin/python -m migrate.nap_khach_cty --ghi    # GHI thật
"""

import json
import pathlib
import re
import sys
import urllib.parse
import urllib.request

from openpyxl import load_workbook

from migrate.doi_chieu_khach_sales import sb, doc_het, khong_dau

ROOT = pathlib.Path(__file__).resolve().parent.parent
FILE = ROOT.parent / "Sales Tracking/import-staging/DM_KHACH_CTY-DAN.xlsx"


def rest(url, key, path, method="GET", body=None, prefer=None):
    p = urllib.parse.quote(path, safe='?&=,.*"<>')
    h = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}
    if prefer:
        h["Prefer"] = prefer
    req = urllib.request.Request(f"{url}/rest/v1/{p}", method=method,
                                 data=json.dumps(body).encode() if body is not None else None, headers=h)
    with urllib.request.urlopen(req) as r:
        return r.read()


def norm_phone(p):
    d = re.sub(r"\D", "", str(p or ""))
    if d.startswith("84"): d = d[2:]
    if d.startswith("0"): d = d[1:]
    return d if len(d) >= 9 else ""


def main():
    ghi = "--ghi" in sys.argv
    url, key = sb()
    cs = doc_het(url, key, "cs_customers?select=id,full_name,primary_phone,channel_id")
    dc = doc_het(url, key, "dim_channel?select=id,channel_l1,channel_l2")

    cs_sdt, cs_ten = {}, {}
    for c in cs:
        pp = norm_phone(c.get("primary_phone"))
        if pp: cs_sdt.setdefault(pp, c)
        cs_ten.setdefault(khong_dau(c.get("full_name")), c)
    dc_theo = {(khong_dau(d["channel_l1"]), khong_dau(d.get("channel_l2"))): d["id"] for d in dc}

    def map_kenh(v):
        s = (v or "").strip()
        if not s or "❓" in s or ";" in s: return None
        parts = [x.strip() for x in s.split("/")]
        l1 = parts[0]; l2 = parts[1] if len(parts) > 1 else ""
        return dc_theo.get((khong_dau(l1), khong_dau(l2)))

    wb = load_workbook(FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip() for h in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}
    def col(r, n):
        i = idx.get(n); return r[i] if i is not None and i < len(r) else None

    gan, them = [], []          # gan: (cs_id, channel_id) · them: dict khách mới
    for r in rows[1:]:
        ten = str(col(r, "Tên khách (nội bộ)") or "").strip()
        cty = str(col(r, "Tên công ty (trên HĐ)") or "").strip()
        if not ten and not cty: continue
        ch = map_kenh(col(r, "Kênh (mới)"))
        sdt = str(col(r, "SĐT liên hệ") or "").strip()
        p = norm_phone(sdt)
        match = cs_sdt.get(p) if p else None
        if not match: match = cs_ten.get(khong_dau(ten))

        if match:
            gan.append((match["id"], ch, match.get("full_name")))
        else:
            mst = str(col(r, "MST") or "").strip()
            ghi_chu = " · ".join(x for x in [
                f"Công ty: {cty}" if cty else "", f"MST: {mst}" if mst else "",
                "nguồn DM_KHACH_CTY"] if x)
            them.append({
                "full_name": ten or cty,
                "primary_phone": sdt or None,
                "address": str(col(r, "Địa chỉ xuất HĐ") or "").strip() or None,
                "needs_phone": not bool(sdt),
                "notes": ghi_chu,
                "source": "DM_KHACH_CTY",
                "trang_thai": "da_duyet",
                "channel_id": ch,
            })

    print(f"{'== GHI ==' if ghi else '== DRY-RUN =='}")
    print(f"Gán channel cho khách đã khớp: {len(gan)} (có channel: {sum(1 for _,c,_ in gan if c)})")
    for cid, ch, ten in gan:
        print(f"  gán {ten}: channel_id={ch}")
    print(f"Import khách mới: {len(them)} (có channel: {sum(1 for t in them if t['channel_id'])})")
    for t in them[:8]:
        print(f"  + {t['full_name']} · sdt={t['primary_phone']} · channel={t['channel_id']}")
    if len(them) > 8: print(f"  … và {len(them)-8} khách nữa")

    if not ghi:
        print("\n(DRY-RUN — thêm --ghi để ghi thật)")
        return

    for cid, ch, _ in gan:
        if ch:
            rest(url, key, f"cs_customers?id=eq.{cid}", method="PATCH",
                 body={"channel_id": ch}, prefer="return=minimal")
    for i in range(0, len(them), 100):
        rest(url, key, "cs_customers", method="POST", body=them[i:i+100], prefer="return=minimal")
    print(f"\n✅ Đã gán {sum(1 for _,c,_ in gan if c)} channel + import {len(them)} khách mới.")


if __name__ == "__main__":
    main()
