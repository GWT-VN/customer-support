"""Quét hợp đồng POE -> gói bảo trì mỗi khách, đối chiếu với file thống kê tay.

NGUỒN SỰ THẬT = HỢP ĐỒNG (user chốt 2026-07-16). File thống kê tay chỉ để đối chiếu.

CÔNG THỨC (suy từ hợp đồng thật, đã kiểm chứng nhiều ca):
    tổng lần bảo trì = số năm × (12 ÷ chu kỳ tháng)
  vd  "01 năm dịch vụ bảo trì định kỳ 5 sao (3 tháng/lần)" -> 1 × 4 = 4 lần
      "01 năm ... (6 tháng/lần)"  -> 1 × 2 = 2  (Chị Nhung Vũ, Anh Tuấn Anh Tây Mỗ)
      "02 năm ... (6 tháng/lần)"  -> 2 × 2 = 4  (Chị Liên)
      "03 năm ... (3 tháng/lần)"  -> 3 × 4 = 12 (Chị Trang Bùi Park City)
      "10 năm ... (3 tháng/lần)"  -> 10 × 4 = 40 (Sixdo = Anh Huy Cận)

⚠️ Hợp đồng MÂU THUẪN NỘI BỘ: nhiều file vừa ghi "Bảo trì bảo dưỡng 4 lần/năm"
   (điều khoản bảo hành, copy cứng) vừa ghi "(6 tháng/lần)" (= 2 lần/năm).
   -> CHỈ tin dòng "N năm dịch vụ bảo trì định kỳ ... (M tháng/lần)"; dòng "4 lần/năm"
      là template, KHÔNG dùng để tính.

⚠️ macOS lưu tên file dạng Unicode NFD -> phải normalize NFC trước khi so tên.

Chạy:  .venv/bin/python -m migrate.quet_hop_dong
"""

import pathlib
import re
import unicodedata
import zipfile
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent
POE_INDEX = {}
POE = ROOT / "Các khách lọc tổng POE"
THONG_KE = ROOT / "Lịch bảo trì - Lịch kĩ thuật/GWT - Lịch bảo trì - Asana.xlsx"
TODAY = date(2026, 7, 17)
OUT = ROOT / f"GWT_goi_bao_tri_tu_hop_dong_{TODAY.isoformat()}.xlsx"
# File bản trước: user đã điền tay cột ĐỐI CHIẾU (tab 1) + Ghi chú (tab 2).
# BẮT BUỘC đọc lại và giữ nguyên — không được để lần chạy sau xoá công của user.
CU = ROOT / "GWT_goi_bao_tri_tu_hop_dong_2026-07-17.xlsx"

SO_CHU = {"một": 1, "hai": 2, "ba": 3, "bốn": 4, "năm": 5, "mười": 10}


def nfc(s):
    return unicodedata.normalize("NFC", str(s or ""))


def khong_dau(s):
    # ⚠️ "đ"/"Đ" (U+0111/U+0110) KHÔNG decompose được bằng NFD -> phải thay tay TRƯỚC,
    #    nếu không "hợp đồng" -> "hop ong" và mọi so khớp tên file đều trượt.
    s = nfc(s).lower().replace("đ", "d")
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]", " ", s)).strip()


# ── Đọc file ─────────────────────────────────────────────────────────────────
def doc_docx(p):
    try:
        with zipfile.ZipFile(p) as z:
            xml = z.read("word/document.xml").decode("utf8", "ignore")
    except Exception:
        return ""
    return re.sub(r"<[^>]+>", "", re.sub(r"</w:p>", "\n", xml))


def doc_xlsx(p):
    try:
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    except Exception:
        return ""
    out = []
    for ws in wb.worksheets:
        for r in ws.iter_rows(values_only=True):
            out += [str(v) for v in r if v is not None and str(v).strip()]
    return "\n".join(out)


def doc_file(p):
    s = p.suffix.lower()
    if s == ".docx":
        return doc_docx(p)
    if s == ".xlsx":
        return doc_xlsx(p)
    return ""            # .doc cũ / .pdf -> không đọc được, sẽ liệt kê riêng


# ── Trích điều khoản bảo trì ─────────────────────────────────────────────────
# Bắt CẢ CỤM trong 1 câu để không nhặt nhầm "6 tháng/lần" của điều khoản khác.
# ⚠️ Số năm phải ĐỨNG SÁT "năm" (chỉ cho phép chú thích trong ngoặc, vd "03 (ba) năm").
#    Bản trước dùng [^)\n]{0,10} tham lam -> với "1 ⇥ 03 năm dịch vụ bảo trì" nó bắt "1"
#    (số thứ tự dòng) rồi nuốt "03" vào phần đệm -> Trang Bùi ra 4 lần thay vì 12.
RE_GOI = re.compile(
    r"(\d{1,2}|một|hai|ba|bốn|năm|mười)\s*(?:\([^)\n]{0,12}\))?\s*năm\s+dịch vụ bảo trì"
    r"[^(\n]{0,40}\(\s*(\d{1,2})\s*tháng\s*/\s*lần\s*\)",
    re.I,
)
RE_NAM_ONLY = re.compile(
    r"(\d{1,2}|một|hai|ba|bốn|năm|mười)\s*(?:\([^)\n]{0,12}\))?\s*năm\s+dịch vụ bảo trì", re.I)
RE_CK_ONLY = re.compile(r"\(\s*(\d{1,2})\s*tháng\s*/\s*lần\s*\)", re.I)
RE_BO = re.compile(r"WH\s*(15|30)\s*A(\s*ECO)?", re.I)

# ── Các mẫu viết KHÁC (quét toàn bộ HĐ mới tìm ra — mẫu chuẩn chỉ chiếm 56/100) ──
# "dịch vụ bảo trì 5 sao 4 lần/1 năm"  (KH anh Tâm)
RE_LAN_TREN_NAM = re.compile(
    r"dịch vụ bảo trì\s*(?:định kỳ\s*)?\d?\s*sao\s*(\d{1,2})\s*lần\s*/\s*(\d{1,2})?\s*năm", re.I)
# "4.2. Tần suất bảo trì: 4 lần/năm"
RE_TAN_SUAT = re.compile(r"tần suất bảo trì\s*:?\s*(\d{1,2})\s*lần\s*/\s*năm", re.I)
# "- Gói bảo trì định kỳ 3 tháng/lần hàng năm cho hệ lọc nước đầu nguồn GE WH30A"
RE_GOI_THANG = re.compile(r"gói bảo trì định kỳ\s*(\d{1,2})\s*tháng\s*/\s*lần\s+hàng năm", re.I)
# Mẫu HĐ MỚI: điều khoản bảo trì nằm ở PHỤ LỤC I, không phải hợp đồng chính
RE_PHU_LUC = re.compile(r"thỏa thuận khác tại Phụ lục", re.I)


def so(v):
    v = str(v).strip().lower()
    return int(v) if v.isdigit() else SO_CHU.get(v)


def trich(txt):
    """-> (số năm, chu kỳ tháng, cách lấy) hoặc (None, None, lý do).

    Thứ tự ưu tiên = độ tin cậy giảm dần. KHÔNG dùng "Bảo trì bảo dưỡng N lần/năm"
    (46 HĐ có dòng này) vì đó là template điều khoản bảo hành copy cứng, mâu thuẫn
    với chu kỳ thật ghi ngay bên cạnh.
    """
    m = RE_GOI.search(txt)                       # "03 năm ... (3 tháng/lần)" — 56 HĐ
    if m:
        return so(m.group(1)), int(m.group(2)), "đủ cụm"

    m = RE_LAN_TREN_NAM.search(txt)              # "dịch vụ bảo trì 5 sao 4 lần/1 năm"
    if m:
        lan, nam = int(m.group(1)), int(m.group(2) or 1)
        if 12 % lan == 0:
            return nam, 12 // lan, "mẫu 'N lần/năm'"

    m = RE_TAN_SUAT.search(txt)                  # "Tần suất bảo trì: 4 lần/năm"
    if m:
        lan = int(m.group(1))
        if 12 % lan == 0:
            n = RE_NAM_ONLY.search(txt)
            return (so(n.group(1)) if n else 1), 12 // lan, "mẫu 'tần suất N lần/năm'"

    m = RE_GOI_THANG.search(txt)                 # "Gói bảo trì định kỳ 3 tháng/lần hàng năm"
    if m:
        n = RE_NAM_ONLY.search(txt)
        return (so(n.group(1)) if n else 1), int(m.group(1)), "mẫu 'gói N tháng/lần hàng năm'"

    n = RE_NAM_ONLY.search(txt)
    c = RE_CK_ONLY.search(txt)
    if n and c:
        return so(n.group(1)), int(c.group(1)), "ghép rời (kiểm tra lại)"
    if n:
        return so(n.group(1)), None, "chỉ có số năm"
    if c:
        return None, int(c.group(1)), "chỉ có chu kỳ"
    if RE_PHU_LUC.search(txt):
        # HĐ mẫu mới đẩy điều khoản sang Phụ lục I -> phải tìm file phụ lục riêng
        return None, None, "HĐ trỏ sang PHỤ LỤC — tìm file phụ lục"
    return None, None, "không thấy điều khoản"


def bo_may(txt, ten_file):
    m = RE_BO.search(ten_file) or RE_BO.search(txt)
    if not m:
        return ""
    return f"WH{m.group(1)}A" + (" ECO" if m.group(2) else "")


# ── Ngày ký hợp đồng ─────────────────────────────────────────────────────────
# Nguồn 1 (tốt nhất, 100 HĐ có): nội dung ghi "ngày 23 tháng 08 năm 2025".
# Nguồn 2 (22 file): số HĐ trong TÊN FILE dạng "001_0424_HĐMB_GWT" -> 0424 = tháng 04/2024.
RE_NGAY = re.compile(r"ngày\s*(\d{1,2})\s*tháng\s*(\d{1,2})\s*năm\s*(\d{4})", re.I)
# ⚠️ CHUẨN hợp đồng VN: ngày ký đứng ngay sau "Hôm nay,". Các ngày khác trong file là
#    ngày phụ lục / căn cứ pháp lý / bàn giao -> lấy "ngày đầu tiên" là SAI.
#    Bằng chứng: 01. Ocean Park — HĐ ghi "Hôm nay, ngày 02 tháng 01 năm 2024" nhưng lấy
#    ngày đầu ra 2025-08-07 (của file khác trong thư mục).
RE_NGAY_KY = re.compile(
    r"Hôm nay,?\s*(?:vào\s*)?ngày\s*(\d{1,2})\s*tháng\s*(\d{1,2})\s*năm\s*(\d{4})", re.I)
RE_SO_HD = re.compile(r"(\d{3})[_/](\d{2})(\d{2})[_/](HĐMB|HDMB|BBBG|ĐNTT|DNTT)", re.I)


def _to_date(m):
    try:
        dt = date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    except ValueError:
        return None
    return dt if 2020 <= dt.year <= TODAY.year else None


def ngay_ky(txt, ten_file):
    """-> (ngày, nguồn). Ưu tiên cụm "Hôm nay, ngày…" = ngày ký thật.
    Không thấy -> KHÔNG đoán, trả ('','') để người tự điền."""
    txt = txt or ""
    m = RE_NGAY_KY.search(txt)
    if m and (dt := _to_date(m)):
        return dt.isoformat(), "ký tại HĐ ('Hôm nay, ngày…')"

    # Fallback 1: số hợp đồng trong TÊN FILE hoặc NỘI DUNG (001/0426/HĐMB -> 04/2026)
    m = RE_SO_HD.search(nfc(ten_file)) or RE_SO_HD.search(txt)
    if m:
        mm, yy = int(m.group(2)), int(m.group(3))
        if 1 <= mm <= 12:
            return f"20{yy:02d}-{mm:02d}", f"suy từ số HĐ {m.group(1)}/{m.group(2)}{m.group(3)} — chỉ tháng/năm"

    # Fallback 2: chỉ khi file có DUY NHẤT 1 ngày -> không thể nhầm
    ngay = [x for x in (_to_date(x) for x in RE_NGAY.finditer(txt)) if x]
    if len(set(ngay)) == 1:
        return ngay[0].isoformat(), "ngày duy nhất trong file (không có 'Hôm nay')"
    if ngay:
        return "", f"⚠️ file có {len(set(ngay))} ngày khác nhau, không có 'Hôm nay' — cần điền tay"
    return "", ""


# ── Giữ ghi chú user từ file bản trước ───────────────────────────────────────
def doc_ghi_chu_cu():
    """-> ({thư mục: đối chiếu}, {thư mục: (năm, ck, tổng, ghi chú)}). File chưa có -> rỗng."""
    dc, dien = {}, {}
    if not CU.exists():
        return dc, dien
    wb = openpyxl.load_workbook(CU, data_only=True)
    if "GÓI BẢO TRÌ TỪ HỢP ĐỒNG" in wb.sheetnames:
        ws = wb["GÓI BẢO TRÌ TỪ HỢP ĐỒNG"]
        hdr = [nfc(c.value) for c in ws[1]]
        if "ĐỐI CHIẾU" in hdr:
            k = hdr.index("ĐỐI CHIẾU")
            for r in ws.iter_rows(min_row=2, values_only=True):
                if r[0] and r[k] and str(r[k]).strip():
                    dc[nfc(r[0])] = nfc(r[k])
    if "THIẾU GÓI - ĐIỀN TAY" in wb.sheetnames:
        ws = wb["THIẾU GÓI - ĐIỀN TAY"]
        hr = next((i for i in range(1, 4)
                   if any(nfc(c.value).startswith("→") for c in ws[i])), 2)
        hdr = [nfc(c.value) for c in ws[hr]]
        idx = {h: i for i, h in enumerate(hdr) if h}
        for r in ws.iter_rows(min_row=hr + 1, values_only=True):
            if not r[0]:
                continue
            g = lambda h: r[idx[h]] if h in idx and idx[h] < len(r) else None
            vals = (g("→ Số năm"), g("→ Chu kỳ (tháng/lần)"), g("→ TỔNG LẦN"), g("→ Ghi chú của bạn"))
            if any(v not in (None, "") for v in vals):
                dien[nfc(r[0])] = vals
    return dc, dien


# ── File thống kê tay (để đối chiếu) ─────────────────────────────────────────
def doc_thong_ke():
    wb = openpyxl.load_workbook(THONG_KE, data_only=True)
    rows = [r for r in wb["Tổng hợp bảo trì"].iter_rows(values_only=True)]
    hdr = [str(h).strip() if h else "" for h in rows[0]]
    i = {h: n for n, h in enumerate(hdr) if h}
    out = []
    for r in rows[1:]:
        ten = r[i["KHÁCH HÀNG"]]
        v = r[i["SỐ LẦN BẢO TRÌ"]]
        if not ten or v is None:
            continue
        # Excel đã đổi "3/4" -> date(2026,4,3): day = đã xong, month = tổng
        if hasattr(v, "month"):
            xong, tong = v.day, v.month
        else:
            m = re.match(r"\s*(\d+)\s*/\s*(\d+)", str(v))
            if not m:
                out.append((nfc(ten), None, None, str(v)))   # "0/10 năm"
                continue
            xong, tong = int(m.group(1)), int(m.group(2))
        out.append((nfc(ten), xong, tong, str(v)))
    return out


def main():
    dc_cu, dien_cu = doc_ghi_chu_cu()   # GIỮ ghi chú user đã điền ở bản trước
    thong_ke = doc_thong_ke()
    tk_index = [(khong_dau(t), t, x, tg, raw) for t, x, tg, raw in thong_ke]

    # ⚠️ CHỈ thư mục KHÁCH = con trực tiếp của "2025/2026 - Khách lẻ lọc tổng" (100 cái).
    #    rglob("*") trả 144 -> gồm cả thư mục con ("01. Chi phí đầu ra"...) -> nhân bản khách.
    khach_dirs = sorted([q for p in POE.iterdir() if p.is_dir()
                         for q in p.iterdir() if q.is_dir()])
    global POE_INDEX
    POE_INDEX = {nfc(d.name): d for d in khach_dirs}

    ket_qua, khong_doc_duoc = [], []
    for d in khach_dirs:
        # file có thể nằm trong thư mục con -> rglob, nhưng chỉ trong phạm vi 1 khách
        files = [p for p in d.rglob("*") if p.is_file() and not p.name.startswith("~$")]
        if not files:
            continue
        # Thứ tự tra user chốt: hợp đồng > phụ lục > biên bản xác nhận > báo giá.
        # (HĐ mẫu mới ghi "thỏa thuận khác tại Phụ lục I" -> gói nằm ở phụ lục.)
        def uu_tien(p):
            n = khong_dau(p.name)
            if "hop dong" in n or "hdmb" in n: return 0
            if "phu luc" in n: return 1
            if "bien ban" in n: return 2
            if "bao gia" in n or "quotation" in n: return 3
            return 4

        nam = ck = None
        cach = nguon = bo = ""
        ngay = ng_nguon = ""
        ngay_tim_duoc = []
        ly_do_cuoi = "không thấy điều khoản"
        # Duyệt HẾT file (không break sớm): HĐ chính có thể trỏ sang phụ lục.
        for p in sorted(files, key=uu_tien):
            txt = doc_file(p)
            if not txt:
                continue
            # Ngày ký: gom TỪ MỌI file hợp đồng (uu_tien 0) rồi mới quyết — 1 thư mục có
            # thể chứa NHIỀU hợp đồng ký ngày khác nhau (mua thêm / đổi máy).
            if uu_tien(p) == 0:
                n_moi, ng_moi = ngay_ky(txt, p.name)
                if n_moi:
                    ngay_tim_duoc.append((n_moi, ng_moi, nfc(p.name)))
            # KHÔNG break sớm: 1 thư mục có thể có nhiều HĐ (mua thêm/đổi máy) — phải
            # duyệt hết để gom đủ ngày ký, nếu không sẽ lấy nhầm ngày của HĐ khác.
            n2, c2, cach2 = trich(txt)
            if n2 and c2 and not (nam and ck):   # gói: lấy cái ĐẦU tiên đủ thông tin
                nam, ck, cach, nguon = n2, c2, cach2, nfc(p.name)
                bo = bo_may(txt, nfc(p.name))
            elif (n2 or c2) and not (nam or ck):  # thiếu 1 vế -> giữ tạm
                nam, ck, cach, nguon = n2, c2, cach2, nfc(p.name)
                bo = bo_may(txt, nfc(p.name))
            elif "PHỤ LỤC" in cach2:
                ly_do_cuoi = cach2
        if not nguon:
            cach = ly_do_cuoi
        # Quyết ngày ký từ các hợp đồng gom được.
        # Nhiều HĐ ngày khác nhau (mua thêm/đổi máy) -> lấy SỚM NHẤT = hợp đồng gốc,
        # nhưng ghi rõ ra để user biết mà kiểm, KHÔNG im lặng chọn hộ.
        ky_that = sorted({x[0] for x in ngay_tim_duoc if "ký tại HĐ" in x[1]})
        if len(ky_that) == 1:
            ngay, ng_nguon = ky_that[0], "ký tại HĐ"
        elif len(ky_that) > 1:
            ngay = ky_that[0]
            ng_nguon = f"⚠️ thư mục có {len(ky_that)} HĐ ký khác ngày ({', '.join(ky_that)}) — lấy sớm nhất"
        elif ngay_tim_duoc:
            ngay, ng_nguon, _ = ngay_tim_duoc[0]
        if not ngay:
            for p in sorted(files, key=uu_tien):
                ngay, ng_nguon = ngay_ky("", p.name)
                if ngay:
                    break
        if not nguon:
            # có file hợp đồng nhưng không đọc được (.doc/.pdf)?
            kho = [nfc(p.name) for p in files
                   if p.suffix.lower() in (".doc", ".pdf") and
                   ("hop dong" in khong_dau(p.name) or "bien ban" in khong_dau(p.name))]
            if kho:
                khong_doc_duoc.append((nfc(d.name), kho))

        tong = nam * (12 // ck) if (nam and ck and 12 % ck == 0) else None

        # khớp với file thống kê theo tên thư mục.
        # ⚠️ Bản trước chỉ cần 2 từ chung -> "Anh Hiếu Park City" khớp nhầm "Chị Trang Bùi
        #    Park City" (chung "park","city"). Nay loại từ ĐỊA DANH/xưng hô rồi mới so,
        #    và bắt buộc trùng >=1 từ TÊN RIÊNG + tỉ lệ trùng cao. Không chắc -> để trống.
        BO_QUA = {"anh", "chi", "co", "chu", "mr", "mrs", "ms", "ong", "ba", "khach",
                  "park", "city", "ecopark", "vinhomes", "vin", "ciputra", "gamuda",
                  "times", "royal", "ocean", "hcm", "hn", "ha", "noi", "eco", "15a", "30a",
                  "wh15a", "wh30a", "poe", "t11", "t12", "q2", "q7", "q9", "q12", "nha"}
        dn = re.sub(r"^\d+\s*", "", khong_dau(d.name))
        b = set(dn.split()) - BO_QUA
        hit, best = None, 0.0
        for tkd, ten, xong, tg, raw in tk_index:
            a = set(tkd.split()) - BO_QUA
            if not a or not b:
                continue
            chung = a & b
            diem = len(chung) / min(len(a), len(b))
            if len(chung) >= 2 and diem >= 0.6 and diem > best:
                hit, best = (ten, xong, tg, raw), diem

        lech = ""
        if hit and tong and hit[2] and tong != hit[2]:
            lech = f"LỆCH: hợp đồng {tong} ≠ thống kê {hit[2]}"
        ket_qua.append([
            nfc(d.name), bo, ngay, ng_nguon, nam, ck, tong,
            hit[0] if hit else "", hit[2] if hit else None, hit[3] if hit else "",
            lech, dc_cu.get(nfc(d.name), ""), cach, nguon,
        ])

    # ── Excel ────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    HEAD = PatternFill("solid", fgColor="1F4E79")
    DO = PatternFill("solid", fgColor="FFC7CE")
    VANG = PatternFill("solid", fgColor="FFEB9C")

    ws = wb.create_sheet("GÓI BẢO TRÌ TỪ HỢP ĐỒNG")
    cols = ["Thư mục hợp đồng", "Bộ", "NGÀY KÝ HĐ", "Nguồn ngày", "Số năm (HĐ)", "Chu kỳ tháng (HĐ)",
            "→ TỔNG LẦN (tính)", "Khách (file thống kê)", "Gói (thống kê)", "Ô gốc", "Máy tự đối chiếu",
            "ĐỐI CHIẾU (ghi chú của bạn)", "Cách lấy", "File nguồn"]
    ws.append(cols)
    for j in range(1, len(cols) + 1):
        c = ws.cell(row=1, column=j)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for r in sorted(ket_qua, key=lambda x: (not x[10], x[0])):   # lệch lên đầu
        ws.append(r)
        i = ws.max_row
        fill = DO if r[10] else (VANG if not r[6] else None)
        for j in range(1, len(cols) + 1):
            cell = ws.cell(row=i, column=j)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill: cell.fill = fill
    for j, w in enumerate([40, 10, 12, 26, 10, 12, 13, 24, 11, 9, 30, 44, 17, 34], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

    # ── Sheet ĐIỀN TAY: khách chưa đọc ra gói + lý do + file có sẵn ───────────
    EDIT = PatternFill("solid", fgColor="DDEBF7")
    ws3 = wb.create_sheet("THIẾU GÓI - ĐIỀN TAY")
    cols3 = ["Thư mục khách (KHÔNG SỬA)", "Vì sao chưa đọc được", "Các file có trong thư mục",
             "→ Số năm", "→ Chu kỳ (tháng/lần)", "→ TỔNG LẦN", "→ Ghi chú của bạn"]
    ws3.append(cols3)
    for j in range(1, len(cols3) + 1):
        c = ws3.cell(row=1, column=j)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEAD if j <= 3 else PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for r in ket_qua:
        if r[6]:
            continue                       # đã có tổng lần -> bỏ qua
        d = POE_INDEX.get(r[0])
        fs = []
        if d:
            fs = sorted({nfc(p.name) for p in d.rglob("*")
                         if p.is_file() and not p.name.startswith("~$")
                         and p.suffix.lower() in (".docx", ".doc", ".pdf", ".xlsx")})
        cu = dien_cu.get(r[0], (None, None, None, None))   # giữ nguyên ô user đã điền
        ws3.append([r[0], r[12] or "không rõ", " · ".join(fs[:8]) or "(không có file tài liệu)",
                    cu[0], cu[1], cu[2], cu[3]])
        i = ws3.max_row
        for j in range(1, len(cols3) + 1):
            cell = ws3.cell(row=i, column=j)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if j >= 4:
                cell.fill = EDIT
    for j, w in enumerate([40, 34, 62, 10, 15, 12, 26], 1):
        ws3.column_dimensions[get_column_letter(j)].width = w
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(cols3))}{ws3.max_row}"
    ws3.insert_rows(1)
    ws3["A1"] = ("CÁCH ĐIỀN: mở file trong cột C, tìm dòng kiểu \"03 năm dịch vụ bảo trì định kỳ 5 sao "
                 "(3 tháng/lần)\" → điền Số năm=3, Chu kỳ=3 → TỔNG LẦN tự tính = năm × (12÷chu kỳ). "
                 "KHÔNG có điều khoản bảo trì thì ghi rõ ở cột Ghi chú.")
    ws3["A1"].font = Font(bold=True, color="9C0006")
    ws3.row_dimensions[1].height = 30

    if khong_doc_duoc:
        ws2 = wb.create_sheet("KẸT ĐỊNH DẠNG")
        ws2.append(["Thư mục", "File (.doc cũ / .pdf — mở ra lưu lại thành .docx là đọc được)"])
        for j in (1, 2):
            c = ws2.cell(row=1, column=j); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD
        for d, fs in khong_doc_duoc:
            ws2.append([d, " · ".join(fs)])
        ws2.column_dimensions["A"].width = 44
        ws2.column_dimensions["B"].width = 80

    wb.save(OUT)

    co_goi = [r for r in ket_qua if r[6]]
    lech = [r for r in ket_qua if r[10]]
    co_ngay = [r for r in ket_qua if r[2]]
    print(f"✓ {OUT.name}\n")
    print(f"  Thư mục khách quét   : {len(ket_qua)}")
    print(f"  Đọc ra gói bảo trì   : {len(co_goi)}")
    print(f"  Không rõ gói         : {len(ket_qua) - len(co_goi)}")
    print(f"  File .doc/.pdf kẹt   : {len(khong_doc_duoc)}")
    print(f"  ⚠️ LỆCH thống kê     : {len(lech)}")
    print(f"  Có NGÀY KÝ hợp đồng  : {len(co_ngay)}")
    print(f"  Giữ ghi chú của user : {len(dc_cu)} (tab 1) + {len(dien_cu)} (tab 2)")
    from collections import Counter
    print("\n  Phân bố gói (theo hợp đồng):")
    for (n, c, t), k in sorted(Counter((r[4], r[5], r[6]) for r in co_goi).items(),
                              key=lambda x: -x[1]):
        print(f"    {k:3} khách: {n} năm × {c} tháng/lần → gói {t} lần")
    if lech:
        print("\n  Các ca LỆCH (user quyết sau):")
        for r in lech:
            print(f"    {r[0][:40]:42} HĐ {r[6]:>2} ≠ thống kê {r[8]:>2}  ({r[7]})")


if __name__ == "__main__":
    main()
