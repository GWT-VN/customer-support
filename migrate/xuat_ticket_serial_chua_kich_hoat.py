"""Xuất Excel các ticket có serial CHƯA kích hoạt bảo hành — để user bổ sung
ngày lắp + khách, rồi kích hoạt sau.

Nguồn: DB project mới (DEST từ migrate/.env.migrate). Không commit file Excel (PII).
Chạy:  .venv/bin/python -m migrate.xuat_ticket_serial_chua_kich_hoat
"""
import json
import pathlib
import urllib.parse
import urllib.request

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent


def dest():
    env = dict(l.strip().split("=", 1) for l in (ROOT / "migrate/.env.migrate").read_text().splitlines()
               if l.strip() and not l.startswith("#") and "=" in l)
    return env["DEST_URL"].strip(), env["DEST_SERVICE_KEY"].strip()


def get(url, key, path):
    p = urllib.parse.quote(path, safe='?&=,.*"')
    req = urllib.request.Request(f"{url}/rest/v1/{p}", headers={"apikey": key, "Authorization": f"Bearer {key}"})
    return json.load(urllib.request.urlopen(req))


def doc_het(url, key, path):
    rows, off = [], 0
    while True:
        b = get(url, key, f"{path}{'&' if '?' in path else '?'}limit=1000&offset={off}")
        rows += b
        if len(b) < 1000:
            return rows
        off += 1000


def main():
    url, key = dest()

    # serial đã kích hoạt (activated) -> loại ra
    activated = {w["serial"] for w in doc_het(url, key, "warranty?select=serial,activated&activated=eq.true")}
    # thông tin serial trong kho (mã nội bộ, tên, khách hiện gắn nếu có)
    kho = {r["serial"]: r for r in doc_het(url, key,
           "v_serial_kho?select=serial,ma_noi_bo,ten_noi_bo,ten_khach,sdt_khach,trang_thai")}
    # khách theo id (cho ticket đã gắn customer_id)
    khach = {c["id"]: c for c in doc_het(url, key, "cs_customers?select=id,full_name,primary_phone")}

    tickets = doc_het(url, key,
        "tickets?select=ticket_code,serial,source_serial,source_customer,customer_id,ticket_type,description,created_at")

    rows = []
    for t in tickets:
        serial = t.get("serial") or t.get("source_serial")
        if not serial or serial in activated:
            continue
        k = kho.get(serial, {})
        c = khach.get(t.get("customer_id")) if t.get("customer_id") else None
        rows.append({
            "ticket": t["ticket_code"],
            "serial": serial,
            "ma_noi_bo": k.get("ma_noi_bo"),
            "ten_may": k.get("ten_noi_bo"),
            "trang_thai": k.get("trang_thai"),
            "khach_ticket": (c or {}).get("full_name") or t.get("source_customer"),
            "sdt_ticket": (c or {}).get("primary_phone"),
            "khach_theo_serial": k.get("ten_khach"),
            "sdt_theo_serial": k.get("sdt_khach"),
            "loai": t.get("ticket_type"),
            "mo_ta": (t.get("description") or "")[:60],
            "ngay_ticket": (t.get("created_at") or "")[:10],
        })
    rows.sort(key=lambda r: r["ticket"], reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ticket serial chưa KH"
    H = ["Mã ticket", "Serial", "Mã nội bộ", "Tên máy", "Trạng thái kho",
         "Khách (ticket)", "SĐT (ticket)", "Khách (theo serial)", "SĐT (theo serial)",
         "Loại", "Mô tả", "Ngày ticket",
         "→ NGÀY LẮP ĐẶT (YYYY-MM-DD)", "→ KHÁCH HÀNG (điền/xác nhận)", "→ SĐT KHÁCH"]
    ws.append(H)
    hf = PatternFill("solid", fgColor="1F4E78"); ef = PatternFill("solid", fgColor="FFF2CC")
    for c in range(1, len(H) + 1):
        cell = ws.cell(1, c); cell.font = Font(bold=True, color="FFFFFF"); cell.fill = hf
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if c > 12:
            cell.fill = PatternFill("solid", fgColor="2E75B6")
    for r in rows:
        ws.append([r["ticket"], r["serial"], r["ma_noi_bo"], r["ten_may"], r["trang_thai"],
                   r["khach_ticket"], r["sdt_ticket"], r["khach_theo_serial"], r["sdt_theo_serial"],
                   r["loai"], r["mo_ta"], r["ngay_ticket"], None, None, None])
        for c in range(13, 16):
            ws.cell(ws.max_row, c).fill = ef
    for c, w in zip(range(1, len(H) + 1), [12, 22, 12, 26, 14, 22, 13, 22, 13, 16, 40, 12, 22, 24, 14]):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(H))}{ws.max_row}"

    out = str(ROOT / "GWT_ticket_serial_chua_kich_hoat_2026-07-29.xlsx")
    wb.save(out)
    print(f"Đã lưu: {out}\n{len(rows)} ticket có serial chưa kích hoạt.")


if __name__ == "__main__":
    main()
