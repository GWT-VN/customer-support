"""Đối chiếu file DM_KHACH_CTY (khách công ty + kênh) ↔ cs_customers + dim_channel.

KHÔNG ghi DB — xuất 1 Excel cho user duyệt trước khi gán channel_id:
  - khớp file -> cs_customers: SĐT liên hệ trước, rồi tên (nội bộ) bỏ dấu
  - map cột "Kênh (mới)" (dạng "channel_l1 / channel_l2") -> dim_channel.id
  - cờ "Cần xem": kênh không map được (❓, đa kênh), hoặc khớp khách bằng tên (mờ)

Chạy:  .venv/bin/python -m migrate.doi_chieu_kenh_cty [--file <xlsx>] [--out <xlsx>]
"""

import pathlib
import re
import sys
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from migrate.doi_chieu_khach_sales import sb, doc_het, khong_dau  # dùng lại helper

ROOT = pathlib.Path(__file__).resolve().parent.parent
FILE_MAC_DINH = ROOT.parent / "Sales Tracking/import-staging/DM_KHACH_CTY-DAN.xlsx"

HEAD = PatternFill("solid", fgColor="1F4E79")
CAM = PatternFill("solid", fgColor="FFEB9C")


def norm_phone(p):
    d = re.sub(r"\D", "", str(p or ""))
    if d.startswith("84"): d = d[2:]
    if d.startswith("0"): d = d[1:]
    return d if len(d) >= 9 else ""


def bang(ws, headers, widths, rows, to_mau=None):
    ws.append(headers)
    for j in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=j)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for r in rows:
        ws.append(r)
        fill = to_mau(r) if to_mau else None
        if fill:
            for j in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=j).fill = fill
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def main():
    fpath = FILE_MAC_DINH
    if "--file" in sys.argv: fpath = pathlib.Path(sys.argv[sys.argv.index("--file") + 1])
    out = ROOT / "migrate/_out/doi_chieu_kenh_cty.xlsx"
    if "--out" in sys.argv: out = pathlib.Path(sys.argv[sys.argv.index("--out") + 1])
    out.parent.mkdir(parents=True, exist_ok=True)

    url, key = sb()
    cs = doc_het(url, key, "cs_customers?select=id,full_name,primary_phone,channel_id")
    dc = doc_het(url, key, "dim_channel?select=id,channel_l1,channel_l2")

    # Chỉ mục cs
    cs_theo_sdt, cs_theo_ten = {}, {}
    for c in cs:
        p = norm_phone(c.get("primary_phone"))
        if p: cs_theo_sdt.setdefault(p, []).append(c)
        cs_theo_ten.setdefault(khong_dau(c.get("full_name")), []).append(c)
    # Chỉ mục dim_channel theo (l1,l2) chuẩn hoá
    dc_theo = {}
    for d in dc:
        dc_theo[(khong_dau(d.get("channel_l1")), khong_dau(d.get("channel_l2")))] = d

    def map_kenh(v):
        s = (v or "").strip()
        if not s or "❓" in s or ";" in s:  # đa kênh / chưa rõ -> để user chọn
            return None, "cần chọn tay" if s else "trống"
        parts = [x.strip() for x in s.split("/")]
        l1 = parts[0]; l2 = parts[1] if len(parts) > 1 else ""
        d = dc_theo.get((khong_dau(l1), khong_dau(l2)))
        if d: return d, f"{d['channel_l1']} / {d.get('channel_l2') or ''}".strip(" /")
        return None, f"KHÔNG map ({s})"

    wb = load_workbook(fpath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip() for h in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}
    def col(r, name):
        i = idx.get(name); return r[i] if i is not None and i < len(r) else None

    out_rows = []
    for r in rows[1:]:
        ten_nb = str(col(r, "Tên khách (nội bộ)") or "").strip()
        if not ten_nb and not (col(r, "Tên công ty (trên HĐ)")): continue
        kenh_file = str(col(r, "Kênh (mới)") or "").strip()
        d, kenh_mo_ta = map_kenh(kenh_file)

        # khớp khách: SĐT -> tên
        p = norm_phone(col(r, "SĐT liên hệ"))
        match, cach = None, "—"
        if p and cs_theo_sdt.get(p):
            match, cach = cs_theo_sdt[p][0], "SĐT"
        else:
            byten = cs_theo_ten.get(khong_dau(ten_nb))
            if byten:
                match, cach = byten[0], "tên (khớp đúng)"
            else:
                cand = [c for c in cs if ten_nb and khong_dau(ten_nb) and khong_dau(ten_nb) in khong_dau(c.get("full_name"))]
                if len(cand) == 1: match, cach = cand[0], "tên (chứa)"
                elif len(cand) > 1: cach = f"tên: {len(cand)} ứng viên"

        can_xem = (d is None) or (match is None) or cach.startswith("tên")
        out_rows.append([
            ten_nb, str(col(r, "Tên công ty (trên HĐ)") or ""), str(col(r, "MST") or ""),
            str(col(r, "Phân loại khách") or ""), kenh_file, kenh_mo_ta,
            (match or {}).get("full_name") or ("— không thấy —" if match is None else ""),
            (match or {}).get("primary_phone") or "", cach,
            "x" if can_xem else "",
        ])

    wb2 = Workbook(); wb2.remove(wb2.active)
    bang(wb2.create_sheet("Đối chiếu kênh"),
         ["Tên khách (file)", "Tên công ty", "MST", "Phân loại", "Kênh (file)",
          "-> dim_channel", "Khớp cs_customers", "SĐT cs", "Cách khớp", "Cần xem"],
         [26, 30, 14, 12, 20, 22, 26, 14, 16, 8], out_rows,
         to_mau=lambda r: CAM if r[9] == "x" else None)
    wb2.save(out)

    n = len(out_rows)
    khop = sum(1 for r in out_rows if not r[6].startswith("—"))
    kenh_ok = sum(1 for r in out_rows if not (r[5].startswith("KHÔNG") or r[5] in ("trống", "cần chọn tay")))
    print(f"File: {n} khách · khớp cs_customers: {khop} · kênh map được: {kenh_ok} · cần xem: {sum(1 for r in out_rows if r[9]=='x')}")
    print(f"-> {out}")


if __name__ == "__main__":
    main()
