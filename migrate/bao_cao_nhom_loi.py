"""Xuất báo cáo nhóm lỗi: bản NỘI BỘ (sếp) + bản GỬI HÃNG (ẩn danh khách).

⚠️ QUYẾT ĐỊNH VỀ RIÊNG TƯ: bản gửi công ty mẹ KHÔNG chứa tên/SĐT/địa chỉ khách —
hãng chỉ cần model + serial + triệu chứng để điều tra lỗi kỹ thuật. Danh tính khách
là dữ liệu của GWT, không có lý do gửi ra ngoài. Bản nội bộ mới có thông tin khách
để CSKH gọi lại.

Chạy:  .venv/bin/python -m migrate.bao_cao_nhom_loi
Ra 3 file:
  GWT_bao_cao_loi_NOIBO_<ngày>.xlsx   — sếp: đủ thông tin, có khách
  GWT_bao_cao_loi_GUI_HANG_<ngày>.xlsx — công ty mẹ: ẩn danh khách
  GWT_bao_cao_loi_TOMTAT_<ngày>.md     — bản đọc nhanh, dán được vào WhatsApp
"""

import json
import pathlib
import urllib.parse
import urllib.request
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent
TODAY = date(2026, 7, 16)

UU_TIEN = {"an_toan": 1, "nghiem_trong": 2, "thuong": 3, "nhe": 4, "khong_loi": 5}
TEN_MUC = {
    "an_toan": "RỦI RO AN TOÀN",
    "nghiem_trong": "Nghiêm trọng",
    "thuong": "Thường",
    "nhe": "Nhẹ",
    "khong_loi": "Không phải lỗi SP",
}


# ── Đọc dữ liệu ──────────────────────────────────────────────────────────────
def sb():
    env = dict(l.strip().split("=", 1) for l in (ROOT / "app-cskh/.env.local").read_text().splitlines()
               if l.strip() and not l.startswith("#") and "=" in l)
    return env["NEXT_PUBLIC_SUPABASE_URL"], env["SUPABASE_SERVICE_ROLE_KEY"]


def get(url, key, path):
    path = urllib.parse.quote(path, safe='?&=,.*"<>')
    r = urllib.request.Request(f"{url}/rest/v1/{path}",
                               headers={"apikey": key, "Authorization": f"Bearer {key}"})
    return json.load(urllib.request.urlopen(r))


def doc_het(url, key, res):
    rows, off = [], 0
    while True:
        sep = "&" if "?" in res else "?"
        batch = get(url, key, f"{res}{sep}limit=1000&offset={off}")
        rows += batch
        if len(batch) < 1000:
            return rows
        off += 1000


# ── Excel helpers ────────────────────────────────────────────────────────────
HEAD = PatternFill("solid", fgColor="1F4E79")
DO = PatternFill("solid", fgColor="FFC7CE")
CAM = PatternFill("solid", fgColor="FFEB9C")


def bang(ws, headers, widths, rows, to_mau=None):
    ws.append(headers)
    for j in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=j)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEAD
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for r in rows:
        ws.append(r)
        i = ws.max_row
        fill = to_mau(r) if to_mau else None
        for j in range(1, len(headers) + 1):
            c = ws.cell(row=i, column=j)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                c.fill = fill
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def main():
    url, key = sb()

    report = [r for r in doc_het(url, key, "v_issue_report?select=*") if r["so_ticket"] > 0]
    report.sort(key=lambda r: (UU_TIEN[r["muc_do"]], -r["so_ticket"]))
    loi_only = [r for r in report if r["muc_do"] != "khong_loi"]
    bao_hang = [r for r in report if r["bao_hang"]]

    ti = doc_het(url, key, "v_ticket_issue?select=*")
    ti.sort(key=lambda r: (UU_TIEN[r["muc_do"]], r["group_code"], r["created_at"]), reverse=False)
    chua_nhom = doc_het(url, key, "v_ticket_chua_phan_nhom?select=*")

    ib = doc_het(url, key, "installed_base?select=internal_code,parent_serial,install_date")
    tickets = doc_het(url, key, "tickets?select=ticket_code,serial,state,created_at")

    # tỷ lệ lỗi theo model — CẢNH BÁO: chỉ tính ticket ĐÃ gắn serial
    so_may = {}
    for m in ib:
        if m["internal_code"]:
            so_may[m["internal_code"]] = so_may.get(m["internal_code"], 0) + 1
    loi_may = {}
    for t in ti:
        if t["muc_do"] != "khong_loi" and t["internal_code"]:
            loi_may.setdefault(t["internal_code"], set()).add(t["ticket_code"])
    ty_le = []
    for code, n in so_may.items():
        if n >= 3:
            k = len(loi_may.get(code, ()))
            ty_le.append([code, n, k, round(100 * k / n, 1)])
    ty_le.sort(key=lambda r: (-r[3], -r[1]))

    ticket_khong_serial = sum(1 for t in tickets if not t["serial"])
    tong_bo = sum(1 for m in ib if not m["parent_serial"])

    # ── 1. Bản NỘI BỘ ────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("TÓM TẮT")
    ws["A1"] = f"BÁO CÁO NHÓM LỖI CSKH — {TODAY.strftime('%d/%m/%Y')} (BẢN NỘI BỘ)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    for line in [
        f"Quy mô: {tong_bo} bộ máy chính / {len(ib)} thiết bị · {len({m['internal_code'] for m in ib})} model · 293 khách",
        f"Ticket: {len(tickets)} tổng · {sum(1 for t in tickets if t['state']=='Open')} đang mở",
        f"Đã gom nhóm: {len({t['ticket_code'] for t in ti})}/{len(tickets)} ticket vào {len(report)} nhóm",
        "",
        "⚠️ ĐỘ TIN CẬY CỦA SỐ LIỆU — đọc trước khi kết luận:",
        f"  · {ticket_khong_serial} ticket CHƯA GẮN SERIAL → không quy được về model."
        "  Tỷ lệ lỗi theo model dưới đây là CẬN DƯỚI, thực tế có thể cao hơn.",
        "  · Nhóm được gom tự động từ mô tả lỗi; ticket thiếu mô tả thì không gom được.",
        "  · Ticket = lượt khách báo, KHÔNG phải số máy hỏng (1 máy có thể báo nhiều lần).",
    ]:
        ws.append([line])
    ws.column_dimensions["A"].width = 120

    bang(wb.create_sheet("NHÓM LỖI"),
         ["Nhóm lỗi", "Mức độ", "Báo hãng?", "Ticket", "Đang mở", "90 ngày qua",
          "Khách", "Máy", "Model dính", "Sớm nhất", "Gần nhất", "Diễn giải"],
         [34, 15, 9, 8, 8, 11, 7, 7, 34, 11, 11, 60],
         [[r["ten"], TEN_MUC[r["muc_do"]], "CÓ" if r["bao_hang"] else "", r["so_ticket"],
           r["dang_mo"], r["trong_90_ngay"], r["so_khach"], r["so_may"],
           r["cac_model"] or "(chưa gắn máy)", r["som_nhat"], r["gan_nhat"], r["mo_ta"]]
          for r in report],
         to_mau=lambda r: DO if r[1] == "RỦI RO AN TOÀN" else (CAM if r[1] == "Nghiêm trọng" else None))

    bang(wb.create_sheet("TỶ LỆ THEO MODEL"),
         ["Model", "Số máy đã lắp", "Ticket lỗi (đã gắn serial)", "Tỷ lệ % (cận dưới)"],
         [22, 14, 24, 18], ty_le,
         to_mau=lambda r: DO if r[3] >= 20 else (CAM if r[3] >= 10 else None))

    bang(wb.create_sheet("TICKET CHI TIẾT"),
         ["Nhóm", "Mức độ", "Ticket", "Ngày", "Trạng thái", "Model", "Serial",
          "Khách", "SĐT", "Tỉnh/TP", "Khách báo gì", "Loại Odoo"],
         [30, 14, 13, 11, 10, 18, 22, 20, 13, 14, 55, 20],
         [[t["nhom_ten"], TEN_MUC[t["muc_do"]], t["ticket_code"], (t["created_at"] or "")[:10],
           t["state"], t["internal_code"] or "", t["serial"] or "", t["customer_name"] or "",
           t["primary_phone"] or "", t["province"] or "", t["description"] or "(không ghi mô tả)",
           t["ticket_type"] or ""] for t in ti],
         to_mau=lambda r: DO if r[1] == "RỦI RO AN TOÀN" else None)

    bang(wb.create_sheet("CHƯA GOM ĐƯỢC"),
         ["Ticket", "Ngày", "Trạng thái", "Loại Odoo", "Mô tả", "Lý do chưa gom"],
         [13, 11, 10, 26, 40, 50],
         [[t["ticket_code"], (t["created_at"] or "")[:10], t["state"], t["ticket_type"] or "",
           t["description"] or "(trống)", t["ly_do"]] for t in chua_nhom])

    f1 = ROOT / f"GWT_bao_cao_loi_NOIBO_{TODAY.isoformat()}.xlsx"
    wb.save(f1)

    # ── 2. Bản GỬI HÃNG — ẩn danh khách ──────────────────────────────────
    wb2 = Workbook()
    wb2.remove(wb2.active)
    ws = wb2.create_sheet("SUMMARY")
    ws["A1"] = f"GWT VIETNAM — FIELD ISSUE REPORT / BÁO CÁO LỖI HIỆN TRƯỜNG · {TODAY.strftime('%d/%m/%Y')}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    for line in [
        f"Installed base: {tong_bo} main units / {len(ib)} devices · {len(so_may)} models",
        f"Tickets analysed: {len(tickets)} (from {min(t['created_at'][:10] for t in tickets)} to {max(t['created_at'][:10] for t in tickets)})",
        "",
        "NOTE ON DATA / LƯU Ý:",
        f"  · {ticket_khong_serial} tickets have no serial attached -> failure rate per model is a LOWER BOUND.",
        "  · Ticket = customer report, not necessarily a distinct failed unit.",
        "  · Customer identity is intentionally excluded / Danh tính khách hàng cố ý không đưa vào.",
    ]:
        ws.append([line])
    ws.column_dimensions["A"].width = 110

    bang(wb2.create_sheet("ISSUE GROUPS"),
         ["Issue group / Nhóm lỗi", "Severity", "Tickets", "Open", "Last 90d",
          "Units affected", "Models", "First seen", "Last seen"],
         [34, 16, 9, 8, 10, 14, 34, 12, 12],
         [[r["ten"], TEN_MUC[r["muc_do"]], r["so_ticket"], r["dang_mo"], r["trong_90_ngay"],
           r["so_may"], r["cac_model"] or "(no serial)", r["som_nhat"], r["gan_nhat"]]
          for r in bao_hang],
         to_mau=lambda r: DO if r[1] == "RỦI RO AN TOÀN" else (CAM if r[1] == "Nghiêm trọng" else None))

    bang(wb2.create_sheet("FAILURE RATE"),
         ["Model", "Units installed", "Tickets with issue", "Rate % (lower bound)"],
         [22, 15, 20, 20], ty_le,
         to_mau=lambda r: DO if r[3] >= 20 else (CAM if r[3] >= 10 else None))

    ma_bh = {r["code"] for r in bao_hang}
    bang(wb2.create_sheet("CASES"),
         ["Issue group", "Severity", "Case ref", "Date", "Status", "Model", "Serial", "Symptom reported"],
         [30, 16, 13, 11, 10, 18, 22, 60],
         [[t["nhom_ten"], TEN_MUC[t["muc_do"]], t["ticket_code"], (t["created_at"] or "")[:10],
           t["state"], t["internal_code"] or "", t["serial"] or "", t["description"] or "(not recorded)"]
          for t in ti if t["group_code"] in ma_bh],
         to_mau=lambda r: DO if r[1] == "RỦI RO AN TOÀN" else None)

    f2 = ROOT / f"GWT_bao_cao_loi_GUI_HANG_{TODAY.isoformat()}.xlsx"
    wb2.save(f2)

    # ── 3. Bản tóm tắt đọc nhanh (dán WhatsApp) ──────────────────────────
    L = [f"*BÁO CÁO LỖI CSKH — {TODAY.strftime('%d/%m/%Y')}*", ""]
    L.append(f"Quy mô: {tong_bo} bộ máy / {len(ib)} thiết bị · {len(tickets)} ticket "
             f"({sum(1 for t in tickets if t['state']=='Open')} đang mở)")
    L.append("")
    an_toan = [r for r in report if r["muc_do"] == "an_toan"]
    if an_toan:
        L.append("🚨 *CẦN QUYẾT NGAY — RỦI RO AN TOÀN*")
        for r in an_toan:
            L.append(f"• *{r['ten']}* — {r['so_ticket']} ca ({r['dang_mo']} đang mở), "
                     f"{r['trong_90_ngay']} ca trong 90 ngày")
            L.append(f"  {r['mo_ta']}")
        L.append("")
    L.append("*Nhóm lỗi lớn nhất*")
    for r in loi_only[:5]:
        xu_huong = " ⬆️ đang tăng" if r["trong_90_ngay"] >= max(1, r["so_ticket"] * 0.7) else ""
        L.append(f"• {r['ten']} — {r['so_ticket']} ca / {r['so_khach']} khách"
                 f" ({r['trong_90_ngay']} trong 90 ngày{xu_huong})")
    L.append("")
    L.append("*Tỷ lệ lỗi cao nhất theo model* (ticket lỗi / máy đã lắp)")
    for code, n, k, pct in ty_le[:5]:
        L.append(f"• {code}: {pct}% ({k}/{n} máy)")
    L.append("")
    L.append("*Việc cần làm*")
    L.append(f"• {ticket_khong_serial} ticket chưa gắn serial → chưa quy được lỗi về model. "
             "Tỷ lệ trên là cận dưới, thực tế có thể cao hơn.")
    thieu_mt = sum(1 for t in chua_nhom if t["ly_do"].startswith("thiếu mô tả"))
    if thieu_mt:
        L.append(f"• {thieu_mt} ticket không ghi mô tả lỗi → không gom nhóm được.")
    L.append(f"• {len(bao_hang)} nhóm đề xuất báo công ty mẹ (file GUI_HANG, đã ẩn danh khách).")

    f3 = ROOT / f"GWT_bao_cao_loi_TOMTAT_{TODAY.isoformat()}.md"
    f3.write_text("\n".join(L), encoding="utf-8")

    print(f"✓ {f1.name}")
    print(f"✓ {f2.name}  (ẩn danh khách)")
    print(f"✓ {f3.name}")
    print()
    print("\n".join(L))


if __name__ == "__main__":
    main()
