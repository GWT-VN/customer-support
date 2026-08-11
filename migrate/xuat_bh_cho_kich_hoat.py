"""Xuất Excel các máy ĐÃ LẮP chờ kích hoạt BH (v_bh_cho_kich_hoat, nguồn da_lap) +
đánh giá logic để user duyệt trước khi ấn kích hoạt hàng loạt.

KHÔNG ghi DB. Mỗi dòng 1 máy: khách/serial/máy/ngày lắp + cờ:
  · OK  (xanh)  = đủ khách + serial + có chính sách BH + có ngày lắp -> ấn được ngay
  · XEM (vàng)  = thiếu ngày lắp (kích hoạt sẽ lấy ngày khác -> hạn có thể sai)
  · KHÔNG (đỏ)  = thiếu khách / serial / chính sách BH -> không kích hoạt tại chỗ

Chạy:  .venv/bin/python -m migrate.xuat_bh_cho_kich_hoat
"""

import pathlib
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from migrate.doi_chieu_khach_sales import sb, doc_het

ROOT = pathlib.Path(__file__).resolve().parent.parent
HEAD = PatternFill("solid", fgColor="1F4E79")
XANH = PatternFill("solid", fgColor="C6EFCE")   # OK
VANG = PatternFill("solid", fgColor="FFEB9C")   # cần xem
DO = PatternFill("solid", fgColor="FFC7CE")     # không kích hoạt được


def bang(ws, headers, widths, rows, to_mau=None):
    ws.append(headers)
    for j in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=j)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for r in rows:
        ws.append(r)
        fill = to_mau(r) if to_mau else None
        if fill:
            for j in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=j).fill = fill
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def main():
    out = ROOT / "migrate/_out/bh_cho_kich_hoat.xlsx"
    if "--out" in sys.argv:
        out = pathlib.Path(sys.argv[sys.argv.index("--out") + 1])
    out.parent.mkdir(parents=True, exist_ok=True)

    url, key = sb()
    v = doc_het(url, key, "v_bh_cho_kich_hoat?nguon=eq.da_lap_chua_kich_hoat"
                "&select=serial,ma_noi_bo,ten_noi_bo,customer_id,ten_khach,sdt_khach,dia_chi,ngay_lap")
    co_cs = {w["internal_code"] for w in doc_het(url, key, "product_warranty?select=internal_code")}

    rows = []
    for r in v:
        thieu = []
        if not r.get("customer_id"): thieu.append("thiếu khách")
        if not (r.get("serial") or "").strip(): thieu.append("thiếu serial")
        if r.get("ma_noi_bo") not in co_cs: thieu.append("không có chính sách BH")
        can_xem = not r.get("ngay_lap")
        if thieu:
            danh_gia, ly_do = "KHÔNG", "; ".join(thieu)
        elif can_xem:
            danh_gia, ly_do = "XEM", "thiếu ngày lắp (sẽ lấy ngày khác)"
        else:
            danh_gia, ly_do = "OK", ""
        rows.append([
            r.get("ten_khach"), r.get("sdt_khach"), r.get("ten_noi_bo"), r.get("ma_noi_bo"),
            r.get("serial"), r.get("ngay_lap"), r.get("dia_chi"), danh_gia, ly_do,
        ])

    rows.sort(key=lambda x: {"OK": 0, "XEM": 1, "KHÔNG": 2}[x[7]])
    mau = {"OK": XANH, "XEM": VANG, "KHÔNG": DO}
    wb = Workbook(); wb.remove(wb.active)
    bang(wb.create_sheet("Chờ kích hoạt (đã lắp)"),
         ["Khách", "SĐT", "Máy", "Mã nội bộ", "Serial", "Ngày lắp", "Địa chỉ", "Đánh giá", "Lý do (nếu chưa OK)"],
         [26, 14, 22, 16, 22, 12, 34, 10, 30], rows,
         to_mau=lambda r: mau[r[7]])
    wb.save(out)

    n = len(rows)
    ok = sum(1 for r in rows if r[7] == "OK")
    xem = sum(1 for r in rows if r[7] == "XEM")
    print(f"Tổng: {n} · OK (ấn được ngay): {ok} · XEM (thiếu ngày): {xem} · KHÔNG: {n-ok-xem}")
    print(f"-> {out}")


if __name__ == "__main__":
    main()
