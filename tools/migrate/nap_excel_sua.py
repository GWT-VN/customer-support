"""Nạp NGƯỢC file Excel soát 3 bảng (user đã điền cột "→") vào Supabase.

NGUYÊN TẮC (user chốt 2026-07-16): CHỈ đẩy data ĐÚNG lên. Mỗi giá trị "→" phải qua
validate; sai/mơ hồ -> TỪ CHỐI kèm lý do, in ra để user sửa vòng sau. Không đoán,
không "sửa giúp" ngoài chuẩn hoá hiển nhiên (bỏ khoảng trắng/chấm trong SĐT).

Đọc theo KHOÁ cột A từng sheet: KHÁCH HÀNG = customers.id · SERIAL-MÁY = serial
· TICKET = ticket_code. Sheet MÁY THIẾU LÕI không ghi DB trực tiếp (product_filter
là masterdata, đổi bằng migration) -> in SQL đề xuất để review.

Chạy:  .venv/bin/python -m migrate.nap_excel_sua "GWT_soat_3_bang_2026-07-16.xlsx"          # dry-run
       .venv/bin/python -m migrate.nap_excel_sua "GWT_soat_3_bang_2026-07-16.xlsx" --ghi    # ghi thật
"""

import json
import pathlib
import re
import sys
import urllib.parse
import urllib.request
from datetime import date, datetime

import openpyxl

ROOT = pathlib.Path(__file__).resolve().parent.parent.parent
TODAY = date(2026, 7, 16)


# ── Validators (thuần, test được) ────────────────────────────────────────────
def normalize_phone(raw):
    """-> (sđt chuẩn '0xxxxxxxxx', lỗi|None). Chỉ chuẩn hoá hiển nhiên, không đoán."""
    if raw is None:
        return None, "rỗng"
    s = str(raw).strip()
    if s.endswith(".0"):                     # Excel đổi số -> float — cắt TRƯỚC khi bỏ dấu chấm
        s = s[:-2]
    s = re.sub(r"[\s.\-()]", "", s)
    if s.startswith("+84"):
        s = "0" + s[3:]
    elif s.startswith("84") and len(s) in (11, 12):
        s = "0" + s[2:]
    if not s.isdigit():
        return None, f"chứa ký tự lạ: “{raw}”"
    if not s.startswith("0") and len(s) in (9, 10):
        s = "0" + s                          # Excel nuốt số 0 đầu
    if len(s) not in (10, 11) or not s.startswith("0"):
        return None, f"không đúng dạng SĐT VN (10-11 số, đầu 0): “{raw}”"
    return s, None


def parse_date_cell(raw):
    """-> (date, lỗi|None). Nhận date/datetime Excel, 'YYYY-MM-DD', 'DD/MM/YYYY'."""
    if raw is None:
        return None, "rỗng"
    if isinstance(raw, datetime):
        return raw.date(), None
    if isinstance(raw, date):
        return raw, None
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date(), None
        except ValueError:
            pass
    return None, f"không đọc được ngày: “{raw}” (dùng YYYY-MM-DD)"


def parse_chu_ky(raw):
    """'12' | '12-24' | '12-24 tháng' -> (min,max, lỗi|None)."""
    if raw is None:
        return None, None, "rỗng"
    s = str(raw).strip()
    m = re.match(r"^(\d+)\s*(?:-\s*(\d+))?(?:\s*tháng)?$", s, re.I)
    if not m:
        return None, None, f"không đọc được chu kỳ: “{raw}” (dùng 12 hoặc 12-24)"
    a = int(m.group(1)); b = int(m.group(2) or a)
    if a <= 0 or b < a:
        return None, None, f"chu kỳ vô lý: “{raw}”"
    return a, b, None


def is_checked(raw):
    return str(raw or "").strip().lower() in ("x", "✔", "✓", "yes", "có", "co", "1", "true")


# ── Supabase REST ────────────────────────────────────────────────────────────
def sb_env():
    env = dict(l.strip().split("=", 1) for l in (ROOT / "apps/web/.env.local").read_text().splitlines()
               if l.strip() and not l.startswith("#") and "=" in l)
    return env["NEXT_PUBLIC_SUPABASE_URL"], env["SUPABASE_SERVICE_ROLE_KEY"]


def req(url, key, path, method="GET", body=None, prefer=None):
    path = urllib.parse.quote(path, safe='?&=,.*"')
    headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}
    if prefer:
        headers["Prefer"] = prefer
    r = urllib.request.Request(f"{url}/rest/v1/{path}", method=method,
                               data=json.dumps(body).encode() if body is not None else None,
                               headers=headers)
    with urllib.request.urlopen(r) as resp:
        raw = resp.read()
        return json.loads(raw) if raw else None


def fetch_all(url, key, resource):
    rows, offset = [], 0
    while True:
        sep = "&" if "?" in resource else "?"
        batch = req(url, key, f"{resource}{sep}limit=1000&offset={offset}")
        rows += batch
        if len(batch) < 1000:
            return rows
        offset += 1000


# ── Đọc sheet theo header ────────────────────────────────────────────────────
def sheet_rows(ws):
    head = [str(c.value).strip() if c.value else "" for c in ws[1]]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None and str(v).strip() for v in r):
            yield dict(zip(head, r))


def val(row, prefix):
    """Lấy giá trị cột '→ <prefix>...' (khớp theo đầu chuỗi, bỏ dấu →)."""
    for k, v in row.items():
        if k.startswith("→") and k[1:].strip().lower().startswith(prefix.lower()):
            if v is not None and str(v).strip() != "":
                return v
    return None


# ── Xử lý từng sheet ─────────────────────────────────────────────────────────
def do_khach(rows, url, key, ok, tuchoi):
    phones = {}          # phone -> customer_id (toàn DB, để bắt trùng)
    for c in fetch_all(url, key, "customers?select=id,primary_phone"):
        if c["primary_phone"]:
            phones[c["primary_phone"]] = c["id"]
    ids = set(phones.values()) | {c["id"] for c in fetch_all(url, key, "customers?select=id")}

    for row in rows:
        cid = str(row.get("id (KHÔNG SỬA)") or "").strip()
        ten_kh = row.get("Tên khách")
        patch, notes = {}, []
        if not cid:
            continue
        if cid not in ids:
            tuchoi.append(("KHÁCH", ten_kh, "id không tồn tại trong DB (cột A bị sửa?)"))
            continue

        sdt = val(row, "SĐT")
        if sdt is not None:
            p, err = normalize_phone(sdt)
            if err:
                tuchoi.append(("KHÁCH", ten_kh, f"SĐT bị từ chối: {err}"))
            elif phones.get(p) not in (None, cid):
                tuchoi.append(("KHÁCH", ten_kh, f"SĐT {p} ĐÃ là SĐT chính của khách khác — cần gộp khách, không tự ghi đè"))
            else:
                patch["primary_phone"] = p
                patch["needs_phone"] = False
                phones[p] = cid
                notes.append(f"SĐT -> {p}")

        t = val(row, "Tên")
        if t is not None:
            if len(str(t).strip()) < 2:
                tuchoi.append(("KHÁCH", ten_kh, f"Tên mới quá ngắn: “{t}”"))
            else:
                patch["full_name"] = str(t).strip()
                notes.append(f"tên -> {patch['full_name']}")

        dc = val(row, "Địa chỉ")
        if dc is not None:
            if len(str(dc).strip()) < 8:
                tuchoi.append(("KHÁCH", ten_kh, f"Địa chỉ mới quá ngắn ({len(str(dc).strip())} ký tự): “{dc}”"))
            else:
                patch["address"] = str(dc).strip()
                notes.append("địa chỉ cập nhật")

        tp = val(row, "Tỉnh")
        if tp is not None:
            patch["province"] = str(tp).strip()
            notes.append(f"tỉnh/TP -> {patch['province']}")

        gc = val(row, "Ghi chú")
        if gc is not None:
            patch["notes"] = str(gc).strip()
            notes.append("ghi chú cập nhật")

        if patch:
            ok.append(("KHÁCH", ten_kh, "; ".join(notes), "customers", f"id=eq.{cid}", patch))


def do_may(rows, url, key, ok, tuchoi):
    catalog = {r["Mã nội bộ"] for r in fetch_all(url, key, 'catalog_item?select="Mã nội bộ"')}
    serials = {r["serial"]: r for r in fetch_all(url, key, "installed_base?select=serial,install_date,internal_code")}

    for row in rows:
        serial = str(row.get("Serial (KHÔNG SỬA)") or "").strip()
        if not serial:
            continue
        if serial not in serials:
            tuchoi.append(("MÁY", serial, "serial không tồn tại trong DB (cột A bị sửa?)"))
            continue
        patch, notes = {}, []

        code = val(row, "Mã nội bộ")
        if code is not None:
            code = str(code).strip()
            if code not in catalog:
                tuchoi.append(("MÁY", serial, f"Mã nội bộ “{code}” KHÔNG có trong catalog — kiểm tra chính tả hoặc tạo mã trước"))
            else:
                patch["internal_code"] = code
                notes.append(f"mã -> {code}")

        d = val(row, "Ngày lắp")
        ngay_lap = None
        if d is not None:
            dt, err = parse_date_cell(d)
            if err:
                tuchoi.append(("MÁY", serial, f"Ngày lắp bị từ chối: {err}"))
            elif dt > TODAY:
                tuchoi.append(("MÁY", serial, f"Ngày lắp ở TƯƠNG LAI: {dt}"))
            else:
                ngay_lap = dt
                patch["install_date"] = dt.isoformat()
                notes.append(f"ngày lắp -> {dt}")

        if patch:
            ok.append(("MÁY", serial, "; ".join(notes), "installed_base", f"serial=eq.{serial}", patch))

        if is_checked(val(row, "Kích hoạt")):
            start = ngay_lap or (parse_date_cell(serials[serial].get("install_date"))[0]
                                 if serials[serial].get("install_date") else None)
            ok.append(("MÁY", serial,
                       f"KÍCH HOẠT BH từ {start or 'hôm nay'}",
                       "__rpc_activate__", serial, {"p_serial": serial,
                                                    **({"p_start": start.isoformat()} if start else {})}))


def do_ticket(rows, url, key, ok, tuchoi):
    serials = {r["serial"] for r in fetch_all(url, key, "installed_base?select=serial")}
    codes = {r["ticket_code"] for r in fetch_all(url, key, "tickets?select=ticket_code")}

    for row in rows:
        tc = str(row.get("Mã ticket (KHÔNG SỬA)") or "").strip()
        if not tc:
            continue
        if tc not in codes:
            tuchoi.append(("TICKET", tc, "mã ticket không tồn tại trong DB (cột A bị sửa?)"))
            continue
        patch, notes = {}, []

        s = val(row, "Serial")
        if s is not None:
            s = str(s).strip()
            if s not in serials:
                tuchoi.append(("TICKET", tc, f"Serial “{s}” KHÔNG có trong installed_base — máy phải vào hệ thống trước"))
            else:
                patch["serial"] = s
                notes.append(f"serial -> {s}")

        st = val(row, "Trạng thái")
        if st is not None:
            st = str(st).strip().capitalize()
            if st not in ("Open", "Done", "Cancel"):
                tuchoi.append(("TICKET", tc, f"Trạng thái “{st}” không hợp lệ (chỉ Open/Done/Cancel)"))
            else:
                patch["state"] = st
                notes.append(f"state -> {st}")

        gc = val(row, "Ghi chú")
        if gc is not None:
            patch["last_note"] = str(gc).strip()
            notes.append("ghi chú xử lý cập nhật")

        if patch:
            ok.append(("TICKET", tc, "; ".join(notes), "tickets", f"ticket_code=eq.{tc}", patch))


def do_thieu_loi(rows, tuchoi):
    """Không ghi DB — sinh SQL đề xuất cho product_filter (masterdata = migration)."""
    sql = []
    for row in rows:
        ma_may = str(row.get("Mã máy (KHÔNG SỬA)") or "").strip()
        loi = val(row, "Mã lõi")
        if not ma_may or loi is None:
            continue
        ten = val(row, "Tên lõi") or ""
        a, b, err = parse_chu_ky(val(row, "Chu kỳ"))
        if err:
            tuchoi.append(("LÕI", f"{ma_may} / {loi}", f"Chu kỳ bị từ chối: {err}"))
            continue
        chu_ky = f"{a} tháng" if a == b else f"{a}-{b} tháng"
        sql.append(f"insert into public.product_filter (\"Máy (model)\", \"Mã lõi lọc\", \"Tên lõi lọc\", \"Chu kỳ thay (tháng)\")\n"
                   f"values ('{ma_may}', '{str(loi).strip()}', '{str(ten).strip()}', '{chu_ky}');")
    return sql


def main():
    args = [a for a in sys.argv[1:] if a != "--ghi"]
    ghi = "--ghi" in sys.argv
    path = ROOT / (args[0] if args else f"GWT_soat_3_bang_{TODAY.isoformat()}.xlsx")
    url, key = sb_env()
    wb = openpyxl.load_workbook(path, data_only=True)

    ok, tuchoi = [], []   # ok: (nhóm, khoá, mô tả, bảng, filter, patch)
    do_khach(list(sheet_rows(wb["KHÁCH HÀNG"])), url, key, ok, tuchoi)
    do_may(list(sheet_rows(wb["SERIAL-MÁY"])), url, key, ok, tuchoi)
    do_ticket(list(sheet_rows(wb["TICKET"])), url, key, ok, tuchoi)
    sql_loi = do_thieu_loi(list(sheet_rows(wb["MÁY THIẾU LÕI"])), tuchoi) if "MÁY THIẾU LÕI" in wb.sheetnames else []

    print(f"═══ SẼ ÁP DỤNG: {len(ok)} thay đổi ═══")
    for nhom, khoa, mota, *_ in ok:
        print(f"  [{nhom}] {khoa}: {mota}")
    print(f"\n═══ TỪ CHỐI: {len(tuchoi)} (sửa lại rồi gửi vòng sau) ═══")
    for nhom, khoa, ly_do in tuchoi:
        print(f"  [{nhom}] {khoa}: {ly_do}")
    if sql_loi:
        print(f"\n═══ SQL ĐỀ XUẤT cho product_filter (masterdata -> cần migration, KHÔNG tự ghi) ═══")
        print("\n".join(sql_loi))

    if not ghi:
        print("\n(dry-run — thêm --ghi để ghi thật)")
        return
    for nhom, khoa, mota, table, flt, patch in ok:
        if table == "__rpc_activate__":
            r = urllib.request.Request(f"{url}/rest/v1/rpc/activate_warranty",
                                       data=json.dumps(patch).encode(), method="POST",
                                       headers={"apikey": key, "Authorization": f"Bearer {key}",
                                                "Content-Type": "application/json"})
            urllib.request.urlopen(r)
        else:
            req(url, key, f"{table}?{flt}", method="PATCH", body=patch, prefer="return=minimal")
    print(f"\nĐÃ GHI {len(ok)} thay đổi vào DB.")


if __name__ == "__main__":
    main()
