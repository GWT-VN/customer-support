"""Đối chiếu khách CSKH (cs_customers) ↔ khách Sales (customers) để user review.

KHÔNG ghi DB — chỉ đọc + xuất 1 file Excel nhiều sheet cho user kiểm tra tay:
  1. "Khớp SĐT"        — cặp cs↔sales trùng SĐT chuẩn hoá (tô cam nếu tên lệch)
  2. "CS chưa khớp"     — khách CSKH không tìm được SĐT bên Sales (nghi alias đại lý)
  3. "Đại lý gộp"       — mọi bản ghi (cả CS lẫn Sales) có tên thuộc nhóm đại lý đồng
                          nghĩa (24 Home/24 Luxury/Hải Nam = 1; CWS = Clean Water Solutions)

Khớp bằng SĐT chuẩn hoá (bỏ ký tự thừa + bỏ 0/84 đầu), KHÔNG khớp bằng tên (bài học
"2 khách tên Yến"). Alias chỉ để GỢI Ý, user tự quyết gộp.

Chạy:  .venv/bin/python -m migrate.doi_chieu_khach_sales
       .venv/bin/python -m migrate.doi_chieu_khach_sales --out /đường/dẫn.xlsx
"""

import json
import pathlib
import re
import sys
import unicodedata
import urllib.parse
import urllib.request

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent.parent

# Nhóm đại lý đồng nghĩa: canonical -> các từ khoá (đã bỏ dấu, thường) để nhận diện.
ALIAS = {
    "Hải Nam (24H)": ["hai nam", "24 home", "24home", "24 luxury", "24luxury", "24h", "24k", "hume"],
    "Clean Water Solutions (CWS)": ["cws", "clean water solution"],
}


def sb():
    env = dict(
        l.strip().split("=", 1)
        for l in (ROOT / "apps/web/.env.local").read_text().splitlines()
        if l.strip() and not l.startswith("#") and "=" in l
    )
    return env["NEXT_PUBLIC_SUPABASE_URL"], env["SUPABASE_SERVICE_ROLE_KEY"]


def get(url, key, path):
    path = urllib.parse.quote(path, safe='?&=,.*"<>')
    r = urllib.request.Request(
        f"{url}/rest/v1/{path}", headers={"apikey": key, "Authorization": f"Bearer {key}"}
    )
    return json.load(urllib.request.urlopen(r))


def doc_het(url, key, res):
    rows, off = [], 0
    while True:
        sep = "&" if "?" in res else "?"
        batch = get(url, key, f"{res}{sep}limit=1000&offset={off}")
        rows += batch
        if len(batch) < 1000:
            return rows
        off += 1000


def khong_dau(s):
    s = unicodedata.normalize("NFD", (s or "").lower())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s.replace("đ", "d")).strip()


def norm_phone(p):
    d = re.sub(r"\D", "", p or "")
    if d.startswith("84"):
        d = d[2:]
    if d.startswith("0"):
        d = d[1:]
    return d if len(d) >= 9 else ""


def nhom_alias(ten):
    kd = khong_dau(ten)
    for canon, tu_khoa in ALIAS.items():
        if any(k in kd for k in tu_khoa):
            return canon
    return ""


# ── Excel ────────────────────────────────────────────────────────────────────
HEAD = PatternFill("solid", fgColor="1F4E79")
CAM = PatternFill("solid", fgColor="FFEB9C")


def bang(ws, headers, widths, rows, to_mau=None):
    ws.append(headers)
    for j in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=j)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEAD
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for r in rows:
        ws.append(r)
        fill = to_mau(r) if to_mau else None
        if fill:
            for j in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=j).fill = fill
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def main():
    out = ROOT / "migrate/_out/doi_chieu_khach_sales.xlsx"
    if "--out" in sys.argv:
        out = pathlib.Path(sys.argv[sys.argv.index("--out") + 1])
    out.parent.mkdir(parents=True, exist_ok=True)

    url, key = sb()
    cs = [c for c in doc_het(url, key,
          "cs_customers?select=id,full_name,primary_phone,address,province,trang_thai")
          if c.get("trang_thai") != "da_xoa"]
    sa = doc_het(url, key,
          "customers?select=id,customer_code,name,phone,phone_chuan,address,province")

    # Chỉ mục Sales theo SĐT chuẩn hoá
    sa_theo_sdt = {}
    for s in sa:
        p = norm_phone(s.get("phone_chuan") or s.get("phone"))
        if p:
            sa_theo_sdt.setdefault(p, []).append(s)

    khop, chua = [], []
    for c in cs:
        p = norm_phone(c.get("primary_phone"))
        matches = sa_theo_sdt.get(p, []) if p else []
        if matches:
            for s in matches:
                ten_lech = khong_dau(c.get("full_name")) != khong_dau(s.get("name"))
                khop.append([
                    c.get("full_name"), c.get("primary_phone"), c.get("province"),
                    s.get("name"), s.get("customer_code"), s.get("phone"), s.get("province"),
                    "LỆCH" if ten_lech else "khớp",
                ])
        else:
            chua.append([
                c.get("full_name"), c.get("primary_phone"), c.get("province"),
                c.get("address"), nhom_alias(c.get("full_name")),
            ])

    # Sheet đại lý gộp: mọi bản ghi (CS + Sales) có tên thuộc nhóm alias
    dai_ly = []
    for c in cs:
        g = nhom_alias(c.get("full_name"))
        if g:
            dai_ly.append([g, "CSKH", c.get("full_name"), c.get("primary_phone"), ""])
    for s in sa:
        g = nhom_alias(s.get("name"))
        if g:
            dai_ly.append([g, "Sales", s.get("name"), s.get("phone"), s.get("customer_code")])
    dai_ly.sort(key=lambda r: (r[0], r[1], r[2] or ""))

    wb = Workbook()
    wb.remove(wb.active)
    bang(wb.create_sheet("Khớp SĐT"),
         ["CS: Tên", "CS: SĐT", "CS: Tỉnh", "Sales: Tên", "Sales: Mã", "Sales: SĐT", "Sales: Tỉnh", "Tên"],
         [26, 14, 16, 26, 12, 14, 16, 8], khop,
         to_mau=lambda r: CAM if r[7] == "LỆCH" else None)
    bang(wb.create_sheet("CS chưa khớp"),
         ["CS: Tên", "CS: SĐT", "CS: Tỉnh", "CS: Địa chỉ", "Nghi đại lý"],
         [26, 14, 16, 40, 22], chua,
         to_mau=lambda r: CAM if r[4] else None)
    bang(wb.create_sheet("Đại lý gộp"),
         ["Nhóm gộp", "Nguồn", "Tên", "SĐT", "Mã Sales"],
         [24, 8, 30, 14, 12], dai_ly)
    wb.save(out)

    print(f"CS (chưa xoá): {len(cs)} · Sales: {len(sa)}")
    print(f"Khớp SĐT: {len(khop)} dòng ({len({r[0] for r in khop})} khách CS) · "
          f"tên lệch: {sum(1 for r in khop if r[7]=='LỆCH')}")
    print(f"CS chưa khớp SĐT: {len(chua)} (nghi đại lý: {sum(1 for r in chua if r[4])})")
    print(f"Bản ghi thuộc nhóm đại lý gộp: {len(dai_ly)}")
    print(f"-> {out}")


if __name__ == "__main__":
    main()
